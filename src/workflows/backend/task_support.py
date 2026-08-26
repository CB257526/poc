"""任务 JSON 序列化、预检与入账口径。"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from difflib import get_close_matches
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from workflows.nodes.node_00_input import Node00Input
from workflows.services import ExcelService

from .config_store import existing_table_file, table_dir
from .errors import ApiError
from .models import Task

TEXT_TYPES = {"图文", "文章", "图文类"}
VIDEO_TYPES = {"视频", "视频类"}


def dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def loads(raw: str | None, default: Any):
    if not raw:
        return default
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return default


def iso(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.isoformat()


def load_media_library() -> tuple[dict[str, str], list[str]]:
    path = existing_table_file("media_library")
    if path is None:
        raise ApiError("尚未配置媒体库", code="VALIDATION_ERROR", status_code=400)
    rows = ExcelService.read_sheet_as_dicts(str(path))
    mapping: dict[str, str] = {}
    names: list[str] = []
    for row in rows:
        name = row.get("媒体") or row.get("媒体名称") or row.get("账号")
        if not name:
            continue
        text = str(name).strip()
        if not text:
            continue
        mapping[Node00Input._normalize_name(text)] = text
        names.append(text)
    mapping.pop("", None)
    # 去重但保持顺序
    seen: set[str] = set()
    ordered = []
    for name in names:
        if name not in seen:
            seen.add(name)
            ordered.append(name)
    return mapping, ordered


def suggest_name(media_name: str, allowed: list[str]) -> str:
    hits = get_close_matches(media_name, allowed, n=1, cutoff=0.45)
    return hits[0] if hits else ""


def first_link(links: Any) -> str:
    if isinstance(links, list):
        return str(links[0]) if links else ""
    return str(links or "")


def preview_validate(input_path: str, corrections: dict[str, str] | None = None) -> dict:
    mapping, allowed = load_media_library()
    rows = ExcelService.read_link_sheet(input_path)
    if not rows:
        raise ApiError("没有读取到有效的媒体与链接", code="VALIDATION_ERROR", status_code=400)

    node = Node00Input()
    dummy_issues: list = []
    merged = node._merge_duplicate_media(rows, dummy_issues)

    records = []
    issues = []
    corrections = {str(k): str(v).strip() for k, v in (corrections or {}).items() if str(v).strip()}

    for idx, row in enumerate(merged):
        record_id = f"rec_{idx + 1:04d}"
        row_number = row.get("row_number")
        media_name = corrections.get(str(row_number), row.get("媒体") or "")
        media_name = str(media_name).strip()
        links = row.get("链接") or []
        matched = bool(media_name) and node._normalize_name(media_name) in mapping
        canonical = mapping.get(node._normalize_name(media_name), media_name) if matched else media_name
        item = {
            "record_id": record_id,
            "row_number": row_number,
            "topic": row.get("主题") or "",
            "media_name": canonical,
            "link_count": len(links) if isinstance(links, list) else (1 if links else 0),
            "link_preview": first_link(links),
            "match_status": "matched" if matched else "unmatched",
            "suggested_name": "" if matched else suggest_name(media_name, allowed),
        }
        records.append(item)
        if not matched:
            issues.append(
                {
                    "record_id": record_id,
                    "node_id": "node_00",
                    "code": "MEDIA_NOT_IN_LIBRARY",
                    "message": "媒体名称无法匹配媒体库",
                    "severity": "error",
                }
            )

    status = "ready" if not issues else "needs_correction"
    return {
        "status": status,
        "allowed_media_names": allowed,
        "records": records,
        "issues": issues,
    }


def workflow_issues(context) -> list[dict]:
    items = []
    for issue in context.issues:
        severity = issue.level if issue.level in {"warning", "error", "critical"} else "error"
        items.append(
            {
                "record_id": issue.record_id,
                "node_id": issue.node_id,
                "code": issue.code,
                "message": issue.message,
                "severity": severity,
            }
        )
    return items


def _fee(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace("¥", "").replace("￥", "").replace(",", "").strip() or 0)
    except (TypeError, ValueError):
        return 0.0


def is_text_type(content_type: str) -> bool:
    return content_type in TEXT_TYPES


def is_video_type(content_type: str) -> bool:
    return content_type in VIDEO_TYPES


def quote_summary_from_context(context) -> dict | None:
    details_raw = (context.quote_details or {}).get("details") or []
    details = []
    text_fee = 0.0
    video_fee = 0.0
    unclassified_fee = 0.0
    total_fee = 0.0
    media_names: set[str] = set()
    for row in details_raw:
        if row.get("eligible_for_monthly_summary") is not True:
            continue
        content_type = str(row.get("文章类型") or "")
        amount = _fee(row.get("费用") or row.get("基础金额"))
        total_fee += amount
        media = str(row.get("媒体") or "")
        media_names.add(media)
        if is_text_type(content_type):
            text_fee += amount
        elif is_video_type(content_type):
            video_fee += amount
        else:
            # 网页抓取失败或类型暂未识别时，费用仍需入账，只暂缓分类。
            unclassified_fee += amount
        details.append(
            {
                "media_name": media,
                "platform": str(row.get("平台") or ""),
                "content_type": content_type,
                "media_level": str(row.get("媒体等级") or ""),
                "followers": str(row.get("粉丝量") or ""),
                "quote_count": 1,
                "unit_price": _fee(row.get("基础金额") or row.get("费用")),
                "amount": amount,
                "status": "完成",
                "title": str(row.get("标题") or ""),
                "publish_url": first_link(row.get("链接")),
                "publish_date": str(row.get("发布日期") or "")[:10],
            }
        )
    if not details:
        return {
            "media_count": 0,
            "quote_count": 0,
            "total_fee": 0,
            "text_fee": 0,
            "video_fee": 0,
            "unclassified_fee": 0,
            "details": [],
        }
    return {
        "media_count": len(media_names),
        "quote_count": len(details),
        "total_fee": round(total_fee, 2),
        "text_fee": round(text_fee, 2),
        "video_fee": round(video_fee, 2),
        "unclassified_fee": round(unclassified_fee, 2),
        "details": details,
    }


def files_payload(task: Task) -> list[dict]:
    items = []
    quote_ready = bool(task.quote_file_path and Path(task.quote_file_path).is_file())
    payment_ready = bool(task.payment_file_path and Path(task.payment_file_path).is_file())
    items.append(
        {
            "key": "quote_detail",
            "filename": Path(task.quote_file_path).name if quote_ready else "2-约稿资料_完成版.xlsx",
            "ready": quote_ready,
        }
    )
    items.append(
        {
            "key": "payment",
            "filename": Path(task.payment_file_path).name if payment_ready else "6-付款.xlsx",
            "ready": payment_ready,
        }
    )
    return items


def progress_payload(task: Task) -> dict:
    data = loads(task.progress_json, None)
    if data:
        return data
    return {"completed_nodes": [], "total_nodes": 7, "current_node": None}


def task_payload(task: Task) -> dict:
    return {
        "task_id": task.id,
        "status": task.status,
        "filename": task.filename,
        "created_at": iso(task.created_at),
        "updated_at": iso(task.updated_at),
        "created_by": task.created_by,
        "error": task.error,
        "progress": progress_payload(task),
        "quote_summary": loads(task.quote_summary_json, None),
        "files": loads(task.files_json, None) or files_payload(task),
        "issues": loads(task.issues_json, []),
    }


def validate_payload(task: Task) -> dict:
    return {
        "task_id": task.id,
        "status": task.status,
        "allowed_media_names": loads(task.allowed_media_json, []),
        "records": loads(task.records_json, []),
        "issues": loads(task.issues_json, []),
    }


def apply_preview(task: Task, preview: dict) -> None:
    task.status = preview["status"]
    task.records_json = dumps(preview["records"])
    task.issues_json = dumps(preview["issues"])
    task.allowed_media_json = dumps(preview["allowed_media_names"])
    task.updated_at = datetime.now()
    task.progress_json = dumps(
        {"completed_nodes": ["node_00"], "total_nodes": 7, "current_node": "node_00"}
        if preview["status"] in {"ready", "needs_correction"}
        else progress_payload(task)
    )


def touch(task: Task) -> None:
    task.updated_at = datetime.now()


STATUS_LABELS = {
    "needs_correction": "待修正",
    "ready": "待处理",
    "running": "处理中",
    "completed": "已完成",
    "failed": "失败",
    "cancelled": "已取消",
}


def parse_month(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    match = re.match(r"(\d{4})[-/](\d{1,2})", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    return text[:7] if len(text) >= 7 else ""


def media_fee_from_quote_sheet(path: str) -> dict[str, float]:
    rows = ExcelService.read_sheet_as_dicts(path, sheet_name="约稿")
    totals: dict[str, float] = defaultdict(float)
    for row in rows:
        media = str(row.get("媒体名称") or row.get("媒体") or "").strip()
        if not media or media == "合计":
            continue
        fee = row.get("基础金额")
        if fee is None:
            fee = row.get("费用")
        totals[media] += _fee(fee)
        totals[media] += _fee(row.get("奖励金额"))
    return dict(totals)


def media_fee_from_summary_sheet(path: str) -> dict[str, float]:
    rows = ExcelService.read_sheet_as_dicts(path, sheet_name="约稿费用合计")
    totals: dict[str, float] = defaultdict(float)
    for row in rows:
        media = str(row.get("媒体名称") or row.get("媒体") or "").strip()
        if not media or media == "合计":
            continue
        fee = row.get("合计费用")
        if fee is None:
            fee = row.get("基础金额")
        totals[media] += _fee(fee)
    return dict(totals)


def write_summary_fees(path: str, summary_fees: dict[str, float]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    if "约稿费用合计" not in wb.sheetnames:
        wb.close()
        raise ApiError("约稿资料缺少「约稿费用合计」表", code="VALIDATION_ERROR", status_code=400)
    ws = wb["约稿费用合计"]
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    try:
        media_col = headers.index("媒体名称") + 1
    except ValueError:
        media_col = 1
    try:
        total_col = headers.index("合计费用") + 1
    except ValueError:
        total_col = 14
    remaining = {k: float(v) for k, v in summary_fees.items()}
    for row_idx in range(2, ws.max_row + 1):
        media = ws.cell(row=row_idx, column=media_col).value
        if not media or str(media).strip() == "合计":
            continue
        name = str(media).strip()
        if name in remaining:
            ws.cell(row=row_idx, column=total_col, value=remaining.pop(name))
    wb.save(path)
    wb.close()
