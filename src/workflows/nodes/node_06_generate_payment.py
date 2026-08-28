"""节点6: 生成约稿资料与云账户付款表"""

from typing import Dict, Any, List, Optional
from datetime import datetime
from collections import defaultdict
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from workflows.nodes.base import BaseNode
from workflows.models import WorkflowContext, NodeOutput, NodeMetrics, Issue
from workflows.services import get_logger

logger = get_logger()

QUOTE_SHEET_HEADERS = [
    "媒体名称", "媒体级别", "粉丝量", "发布形式", "约稿类型",
    "平台", "标题", "发布链接", "作品截图", "发布日期",
    "约稿数量", "户名", "身份证", "账号", "电话",
    "开户行", "开户行所在城市", "基础金额", "奖励金额", "同步平台",
]

FEE_SUMMARY_HEADERS = [
    "媒体名称", "媒体级别", "发布形式", "约稿类型", "约稿数量",
    "户名", "身份证", "账号", "电话", "开户行",
    "开户行所在城市", "基础金额", "奖励金额", "合计费用",
]

PAYMENT_TEMPLATE_ID = "TEMPLATE-BANK-YZH006"
PAYMENT_TEMPLATE_HINT = (
    "自助签约发起付款表格模板-银行卡(单个批次文件最大支持10000条订单,"
    "单笔订单最大支持金额可在『业务中心-合作信息』中查看,请勿修改此条信息)"
)
PAYMENT_FORMULA_LAST_ROW = 9420

# 作品截图：单元格里只缩小显示尺寸，xlsx 里仍嵌入原图像素，避免压糊。
SCREENSHOT_DISPLAY_HEIGHT_PX = 180
SCREENSHOT_ROW_HEIGHT_PT = 135
SCREENSHOT_COLUMN_WIDTH = 28


class Node06GeneratePayment(BaseNode):
    """
    生成约稿资料与付款表节点

    职责：
    1. 按收款方汇总费用，生成云账户银行卡上传模板
    2. 写出与 table/2-约稿资料.xlsx 对齐的约稿资料（约稿 + 约稿费用合计）
    3. 保存输出文件路径
    """

    def __init__(self):
        super().__init__("node_06", "生成付款表")

    def process(self, context: WorkflowContext) -> NodeOutput:
        """处理付款表生成"""
        start_time = datetime.now()
        metrics = NodeMetrics()
        issues = []

        # 检查是否有约稿明细数据
        if not context.quote_details or not context.quote_details.get("details"):
            issue = Issue(
                level="critical",
                code="NO_QUOTE_DETAILS",
                message="没有约稿明细数据，无法生成付款表",
                node_id=self.node_id
            )
            issues.append(issue)
            return NodeOutput.create_failure(metrics=metrics, issues=issues)

        # 防御性二次过滤：即使上游或历史数据意外混入异常明细，也不进入
        # 月度汇总、付款文件或约稿输出。
        quote_details = [
            detail for detail in context.quote_details["details"]
            if detail.get("eligible_for_monthly_summary") is True
        ]
        if not quote_details:
            issue = Issue(
                level="critical",
                code="NO_ELIGIBLE_QUOTE_DETAILS",
                message="没有校验通过的约稿明细，异常数据不会计入月度汇总",
                node_id=self.node_id,
            )
            return NodeOutput.create_failure(metrics=metrics, issues=[issue])
        metrics.processed_count = len(quote_details)

        try:
            # 1. 生成月度汇总
            monthly_summary = self._generate_monthly_summary(quote_details)
            context.monthly_summary = monthly_summary

            logger.info(
                "monthly_summary_generated",
                months=len(monthly_summary)
            )

            # 2. 生成付款行数据
            payment_rows = self._generate_payment_rows(quote_details)
            context.payment_rows = payment_rows

            logger.info(
                "payment_rows_generated",
                count=len(payment_rows)
            )

            # 3. 写入云账户付款上传模板（按 run_id 分子目录）
            from workflows.paths import run_output_dir

            output_dir = context.config.get("output_dir") or str(
                run_output_dir(context.run_id, context.config.get("output_root"))
            )
            output_file = self._write_payment_excel(context, payment_rows, output_dir=output_dir)
            context.output_files["payment"] = output_file

            logger.info(
                "payment_excel_generated",
                output_file=output_file
            )

            # 4. 写入完善后的约稿资料表
            quote_file = self._write_quote_detail_excel(context, quote_details, output_dir=output_dir)
            context.output_files["quote_detail"] = quote_file

            logger.info(
                "quote_detail_excel_generated",
                output_file=quote_file
            )

            metrics.success_count = len(quote_details)

        except Exception as e:
            issue = Issue(
                level="critical",
                code="GENERATION_FAILED",
                message=f"生成付款表失败: {str(e)}",
                node_id=self.node_id
            )
            issues.append(issue)
            metrics.error_count = len(quote_details)
            logger.error(
                "payment_generation_failed",
                error=str(e)
            )
            return NodeOutput.create_failure(metrics=metrics, issues=issues)

        # 计算耗时
        duration = (datetime.now() - start_time).total_seconds() * 1000
        metrics.duration_ms = duration

        logger.info(
            "node_completed",
            node_id=self.node_id,
            processed=metrics.processed_count,
            success=metrics.success_count,
            output_file=context.output_files.get("payment"),
            duration_ms=duration
        )

        return NodeOutput.create_success(
            metrics=metrics,
            issues=issues,
            data={
                "monthly_summary": monthly_summary,
                "payment_rows": payment_rows
            }
        )

    def _generate_monthly_summary(self, details: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        生成月度汇总

        Args:
            details: 约稿明细列表

        Returns:
            月度汇总字典 {月份: {媒体数, 文章数, 总费用}}
        """
        monthly_data = defaultdict(lambda: {
            "media_count": set(),
            "article_count": 0,
            "total_fee": 0.0
        })

        for detail in details:
            if detail.get("eligible_for_monthly_summary") is not True:
                continue
            publish_date = detail.get("发布日期")
            media = detail.get("媒体")
            fee = detail.get("费用", 0)

            # 解析月份
            month = self._extract_month(publish_date)
            if not month:
                continue

            # 累计数据
            monthly_data[month]["media_count"].add(media)
            monthly_data[month]["article_count"] += 1
            monthly_data[month]["total_fee"] += float(fee) if fee else 0.0

        # 转换为最终格式
        summary = {}
        for month, data in monthly_data.items():
            summary[month] = {
                "媒体数": len(data["media_count"]),
                "文章数": data["article_count"],
                "总费用": data["total_fee"]
            }

        return summary

    def _generate_payment_rows(self, details: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        生成付款行数据（按收款方汇总）

        Args:
            details: 约稿明细列表

        Returns:
            付款行列表
        """
        # 按收款方分组
        payee_groups = defaultdict(lambda: {
            "details": [],
            "total_fee": 0.0
        })

        for detail in details:
            payee = detail.get("收款方")
            if not payee:
                continue

            fee = detail.get("费用", 0)
            payee_groups[payee]["details"].append(detail)
            payee_groups[payee]["total_fee"] += float(fee) if fee else 0.0

        # 生成付款行
        payment_rows = []
        for payee, group_data in payee_groups.items():
            # 获取账户信息（从第一条记录）
            first_detail = group_data["details"][0]

            row = {
                "收款方": payee,
                "户名": payee,
                "开户行": first_detail.get("开户行"),
                "账号": first_detail.get("账号"),
                "身份证": first_detail.get("身份证"),
                "联系方式": first_detail.get("联系方式"),
                "电话": first_detail.get("联系方式"),
                "文章数": len(group_data["details"]),
                "总费用": group_data["total_fee"],
                "基础服务费": group_data["total_fee"],
                "备注": "",
            }
            payment_rows.append(row)

        # 按约稿资料中首次出现顺序，与样表 6-付款 一致
        return payment_rows

    def _write_payment_excel(
        self,
        context: WorkflowContext,
        payment_rows: List[Dict[str, Any]],
        output_dir: str = "./output",
    ) -> str:
        """写入云账户银行卡付款上传模板（对齐 table/6-付款.xlsx）。"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"付款表_{timestamp}.xlsx"
        batch_name = os.path.splitext(output_filename)[0]
        output_path = os.path.join(output_dir, output_filename)
        os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "上传模板"

        header_font = Font(name="微软雅黑", size=12)
        header_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")

        ws["A1"] = PAYMENT_TEMPLATE_ID
        ws["A1"].font = header_font
        ws["B1"] = PAYMENT_TEMPLATE_HINT
        ws["B1"].font = header_font
        ws.merge_cells("B1:F1")

        ws["A2"] = "批次号(与文件名称一致，必填)"
        ws["B2"] = "总笔数"
        ws["C2"] = "总金额/元(以显示金额汇总)"
        for col in range(1, 4):
            ws.cell(row=2, column=col).font = header_font

        ws["A3"] = batch_name
        ws["B3"] = f"=COUNTA(F5:F{PAYMENT_FORMULA_LAST_ROW})"
        ws["C3"] = f"=SUM(F5:F{PAYMENT_FORMULA_LAST_ROW})"
        for col in range(1, 4):
            ws.cell(row=3, column=col).font = header_font

        column_labels = [
            "平台企业订单号(非必填)",
            "收款账号(个人银行卡号,必填)",
            "收款户名(真实姓名,必填)",
            "身份证号(必填)",
            "联系电话(签约手机号,必填)",
            "基础服务费金额/元(四舍五入至分,必填)",
            "备注(36个字符以内,非必填)",
        ]
        for col_idx, label in enumerate(column_labels, start=1):
            cell = ws.cell(row=4, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row_data in enumerate(payment_rows, start=5):
            ws.cell(row=row_idx, column=2, value=row_data.get("账号"))
            ws.cell(row=row_idx, column=3, value=row_data.get("户名") or row_data.get("收款方"))
            ws.cell(row=row_idx, column=4, value=row_data.get("身份证"))
            ws.cell(row=row_idx, column=5, value=row_data.get("电话") or row_data.get("联系方式"))
            fee = row_data.get("基础服务费")
            if fee is None:
                fee = row_data.get("总费用")
            ws.cell(row=row_idx, column=6, value=fee)
            ws.cell(row=row_idx, column=7, value=row_data.get("备注") or None)
            ws.row_dimensions[row_idx].height = 22

        ws.column_dimensions["A"].width = 22.38
        ws.column_dimensions["B"].width = 41.54
        ws.column_dimensions["C"].width = 14.38
        ws.column_dimensions["D"].width = 21.62
        ws.column_dimensions["E"].width = 18.54
        ws.column_dimensions["F"].width = 11.92
        ws.column_dimensions["G"].width = 29.5

        wb.save(output_path)
        return output_path

    @staticmethod
    def _set_cell_value(ws, row_idx: int, col_idx: int, value: Any) -> None:
        """
        写入单元格值；以 = 开头的字符串（如 WPS 的 =DISPIMG(...) 图片引用）
        强制按文本存储，避免 openpyxl 当作公式导致读回为空。
        """
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(value, str) and value.startswith("="):
            cell.value = value
            cell.data_type = "s"
        else:
            cell.value = value

    @staticmethod
    def _extract_month(date_str: Any) -> str:
        """
        从日期字符串中提取月份

        Args:
            date_str: 日期字符串（如 "2024-08-15" 或 "2024/8/15"）

        Returns:
            月份字符串（如 "2024-08"），失败返回空字符串
        """
        if not date_str:
            return ""

        try:
            # 尝试多种日期格式
            date_str = str(date_str).strip()

            # 处理 datetime 对象
            if isinstance(date_str, datetime):
                return date_str.strftime("%Y-%m")

            # 处理字符串
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m", "%Y/%m"]:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime("%Y-%m")
                except ValueError:
                    continue

            # 如果上述都失败，尝试提取前7个字符（YYYY-MM）
            if len(date_str) >= 7 and date_str[4] in ['-', '/']:
                return date_str[:7]

        except Exception:
            pass

        return ""

    def _write_quote_detail_excel(
        self,
        context: WorkflowContext,
        quote_details: List[Dict[str, Any]],
        output_dir: str = "./output",
    ) -> str:
        """写入约稿资料（对齐 table/2-约稿资料.xlsx：约稿 + 约稿费用合计）。"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"2-约稿资料_完善版_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "约稿"

        for col_idx, header in enumerate(QUOTE_SHEET_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.column_dimensions["I"].width = SCREENSHOT_COLUMN_WIDTH

        last_data_row = 1
        for row_idx, detail in enumerate(quote_details, start=2):
            last_data_row = row_idx
            base_amount = detail.get("基础金额")
            if base_amount is None:
                base_amount = detail.get("费用")

            ws.cell(row=row_idx, column=1, value=detail.get("媒体"))
            ws.cell(row=row_idx, column=2, value=detail.get("媒体等级"))
            ws.cell(row=row_idx, column=3, value=detail.get("粉丝量"))
            # 发布形式仍由业务人员下载后手动填写。
            ws.cell(row=row_idx, column=4, value=None)
            ws.cell(row=row_idx, column=5, value=detail.get("文章类型"))
            ws.cell(row=row_idx, column=6, value=detail.get("平台"))
            ws.cell(row=row_idx, column=7, value=detail.get("标题"))
            ws.cell(row=row_idx, column=8, value=detail.get("链接"))
            self._embed_screenshot(ws, row_idx, detail.get("截图"))
            ws.cell(row=row_idx, column=10, value=detail.get("发布日期"))
            ws.cell(row=row_idx, column=11, value=1)
            ws.cell(row=row_idx, column=12, value=detail.get("收款方"))
            ws.cell(row=row_idx, column=13, value=detail.get("身份证"))
            ws.cell(row=row_idx, column=14, value=detail.get("账号"))
            ws.cell(row=row_idx, column=15, value=detail.get("联系方式"))
            ws.cell(row=row_idx, column=16, value=detail.get("开户行"))
            ws.cell(row=row_idx, column=17, value=detail.get("开户行所在城市"))
            ws.cell(row=row_idx, column=18, value=base_amount)
            ws.cell(row=row_idx, column=19, value=detail.get("奖励金额"))
            ws.cell(row=row_idx, column=20, value=detail.get("同步平台"))

        footer_row = last_data_row + 1
        if last_data_row >= 2:
            ws.cell(row=footer_row, column=17, value="合计")
            ws.cell(row=footer_row, column=18, value=f"=SUM(R2:R{last_data_row})")

        ws_sum = wb.create_sheet("约稿费用合计")
        for col_idx, header in enumerate(FEE_SUMMARY_HEADERS, start=1):
            ws_sum.cell(row=1, column=col_idx, value=header)

        for row_idx, detail in enumerate(quote_details, start=2):
            base_amount = detail.get("基础金额")
            if base_amount is None:
                base_amount = detail.get("费用")
            reward = detail.get("奖励金额") or 0
            try:
                total_fee = float(base_amount or 0) + float(reward or 0)
            except (TypeError, ValueError):
                total_fee = base_amount

            # 约稿费用合计中这两列与「约稿」表对调：发布形式=视频/图文，约稿类型=原创/通稿
            ws_sum.cell(row=row_idx, column=1, value=detail.get("媒体"))
            ws_sum.cell(row=row_idx, column=2, value=detail.get("媒体等级"))
            ws_sum.cell(row=row_idx, column=3, value=detail.get("文章类型"))
            ws_sum.cell(row=row_idx, column=4, value=None)
            ws_sum.cell(row=row_idx, column=5, value=1)
            ws_sum.cell(row=row_idx, column=6, value=detail.get("收款方"))
            ws_sum.cell(row=row_idx, column=7, value=detail.get("身份证"))
            ws_sum.cell(row=row_idx, column=8, value=detail.get("账号"))
            ws_sum.cell(row=row_idx, column=9, value=detail.get("联系方式"))
            ws_sum.cell(row=row_idx, column=10, value=detail.get("开户行"))
            ws_sum.cell(row=row_idx, column=11, value=detail.get("开户行所在城市"))
            ws_sum.cell(row=row_idx, column=12, value=base_amount)
            ws_sum.cell(row=row_idx, column=13, value=detail.get("奖励金额"))
            ws_sum.cell(row=row_idx, column=14, value=total_fee)

        self._merge_payee_totals(ws_sum, quote_details)

        if last_data_row >= 2:
            ws_sum.cell(row=footer_row, column=11, value="合计")
            ws_sum.cell(row=footer_row, column=12, value=f"=SUM(L2:L{last_data_row})")
            ws_sum.cell(row=footer_row, column=14, value=f"=SUM(N2:N{last_data_row})")

        wb.save(output_path)
        return output_path

    def _embed_screenshot(self, ws, row_idx: int, screenshot_path: Optional[str]) -> None:
        if screenshot_path and os.path.exists(screenshot_path):
            try:
                from openpyxl.drawing.image import Image
                from PIL import Image as PILImage

                with PILImage.open(screenshot_path) as original_img:
                    orig_w, orig_h = original_img.size
                img = Image(screenshot_path)
                if orig_h:
                    img.height = SCREENSHOT_DISPLAY_HEIGHT_PX
                    img.width = max(1, int(SCREENSHOT_DISPLAY_HEIGHT_PX * orig_w / orig_h))
                img.anchor = f"I{row_idx}"
                ws.add_image(img)
                ws.row_dimensions[row_idx].height = SCREENSHOT_ROW_HEIGHT_PT
                return
            except Exception as e:
                logger.warning("screenshot_embed_failed", path=screenshot_path, error=str(e))
                ws.cell(row=row_idx, column=9, value=screenshot_path)
                return
        ws.cell(row=row_idx, column=9, value="")

    @staticmethod
    def _merge_payee_totals(ws, quote_details: List[Dict[str, Any]]) -> None:
        """同一户名连续行合并「合计费用」，金额写在合并区第一格。"""
        if not quote_details:
            return

        def payee_of(detail: Dict[str, Any]) -> str:
            return str(detail.get("收款方") or "")

        def amount_of(detail: Dict[str, Any]) -> float:
            base = detail.get("基础金额")
            if base is None:
                base = detail.get("费用") or 0
            reward = detail.get("奖励金额") or 0
            try:
                return float(base or 0) + float(reward or 0)
            except (TypeError, ValueError):
                return 0.0

        start = 0
        while start < len(quote_details):
            end = start
            current = payee_of(quote_details[start])
            total = amount_of(quote_details[start])
            while end + 1 < len(quote_details) and payee_of(quote_details[end + 1]) == current and current:
                end += 1
                total += amount_of(quote_details[end])

            first_excel_row = start + 2
            last_excel_row = end + 2
            ws.cell(row=first_excel_row, column=14, value=total)
            for row_idx in range(first_excel_row + 1, last_excel_row + 1):
                ws.cell(row=row_idx, column=14, value=None)
            if last_excel_row > first_excel_row:
                ws.merge_cells(
                    start_row=first_excel_row,
                    start_column=14,
                    end_row=last_excel_row,
                    end_column=14,
                )
            start = end + 1
