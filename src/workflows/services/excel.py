"""Excel读写服务"""

from openpyxl import load_workbook, Workbook
from typing import List, Dict, Any, Optional
from pathlib import Path
import structlog

logger = structlog.get_logger()


class ExcelService:
    """Excel文件读写服务"""

    @staticmethod
    def _clean_cell(value: Any) -> Optional[str]:
        """清洗单元格值：空值返回 None，否则去除首尾空白并转为字符串"""
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    @staticmethod
    def read_link_sheet(
        file_path: str,
        sheet_name: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """
        读取无表头的"主题-媒体-链接"分层表（如 1-链接.xlsx）

        结构约定（该表无表头，共两列）：
        - A列是分层标签，B列是对应链接列表（一对多）
        - A列非空且B列为空：主题行（如 "主题1"），后续链接归属于该主题
        - A列非空且B列非空：媒体（作者）块起点，A列通常为合并单元格、仅左上角有值
        - B列非空：该媒体名下的一条链接（一行一条）
        - 全空行：主题间的分隔，跳过

        Returns:
            [{"主题": str, "媒体": str, "row_number": int, "链接": [str, ...]}, ...]
            每条记录对应一个媒体（作者）块，链接为该媒体名下的所有链接列表
            row_number 为媒体块起始行的 Excel 行号（1-based）
        """
        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active

            records: List[Dict[str, Any]] = []
            current_topic: Optional[str] = None
            current_media: Optional[str] = None
            media_start_row: Optional[int] = None
            media_links: List[str] = []

            def flush_media() -> None:
                """把已聚合的媒体块写入结果"""
                nonlocal current_media, media_start_row, media_links
                if current_media and media_links:
                    records.append({
                        "主题": current_topic,
                        "媒体": current_media,
                        "row_number": media_start_row,
                        "链接": media_links,
                    })
                current_media = None
                media_start_row = None
                media_links = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                topic_val = ExcelService._clean_cell(row[0] if len(row) >= 1 else None)
                link_val = ExcelService._clean_cell(row[1] if len(row) >= 2 else None)

                if not link_val:
                    # 主题行或空行：结束当前媒体块，更新主题
                    if topic_val:
                        flush_media()
                        current_topic = topic_val
                    else:
                        flush_media()
                    continue

                # B列有链接：若本行A列非空，则是新媒体（作者）块起点
                if topic_val:
                    flush_media()
                    current_media = topic_val
                    media_start_row = row_idx

                if not current_media:
                    # 主题行之前的孤儿链接，跳过并告警
                    logger.warning(
                        "link_without_media",
                        file_path=file_path,
                        row=row_idx,
                    )
                    continue

                media_links.append(link_val)

            # 收尾：写入最后一个媒体块
            flush_media()

            logger.info(
                "link_sheet_read_success",
                file_path=file_path,
                sheet_name=sheet_name or ws.title,
                medias_read=len(records),
                links_read=sum(len(r["链接"]) for r in records),
            )
            return records
        finally:
            wb.close()

    @staticmethod
    def read_sheet_as_dicts(
        file_path: str,
        sheet_name: Optional[str] = None,
        skip_rows: int = 0,
        max_rows: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """
        读取Excel sheet并转换为字典列表

        Args:
        
            file_path: Excel文件路径
            sheet_name: Sheet名称，None表示第一个sheet
            skip_rows: 跳过的行数（通常用于跳过标题前的说明）
            max_rows: 最多读取的行数，None表示全部读取

        Returns:
            字典列表，每个字典代表一行数据
        """
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            # 获取表头（第一行数据）
            # 注意：空表头用占位名填充，避免生成 None 键导致下游 get("字段") 全部落空
            headers = []
            header_row = skip_rows + 1
            for col_idx, cell in enumerate(ws[header_row], start=1):
                value = cell.value
                if value is None or str(value).strip() == "":
                    headers.append(f"_col_{col_idx}")
                else:
                    headers.append(str(value).strip())

            # 读取数据
            rows = []
            start_row = skip_rows + 2
            end_row = ws.max_row + 1

            if max_rows:
                end_row = min(end_row, start_row + max_rows)

            for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
                # 跳过空行
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                row_dict = {}
                for header, value in zip(headers, row):
                    # 处理None值
                    if value is None:
                        row_dict[header] = None
                    else:
                        # 转换为字符串并去除首尾空格
                        row_dict[header] = str(value).strip() if value != "" else None

                rows.append(row_dict)

            wb.close()

            logger.info(
                "excel_read_success",
                file_path=file_path,
                sheet_name=sheet_name or ws.title,
                rows_read=len(rows),
                headers=headers
            )

            return rows

        except Exception as e:
            logger.error(
                "excel_read_failed",
                file_path=file_path,
                sheet_name=sheet_name,
                error=str(e)
            )
            raise

    @staticmethod
    def get_metadata(file_path: str) -> Dict[str, Any]:
        """
        获取Excel文件元信息

        Args:
            file_path: Excel文件路径

        Returns:
            包含文件大小、sheet名称、行列数等信息的字典
        """
        import os

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)

            metadata = {
                "file_path": file_path,
                "size_bytes": os.path.getsize(file_path),
                "sheet_names": wb.sheetnames,
                "sheets": {}
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                metadata["sheets"][sheet_name] = {
                    "max_row": ws.max_row,
                    "max_column": ws.max_column
                }

            logger.info(
                "excel_metadata_extracted",
                file_path=file_path,
                sheets=len(wb.sheetnames)
            )

            wb.close()

            return metadata

        except Exception as e:
            logger.error(
                "excel_metadata_failed",
                file_path=file_path,
                error=str(e)
            )
            raise

    @staticmethod
    def write_dicts_to_excel(
        data: List[Dict[str, Any]],
        output_path: str,
        sheet_name: str = "Sheet1",
        headers: Optional[List[str]] = None
    ) -> str:
        """
        将字典列表写入Excel文件

        Args:
            data: 数据列表
            output_path: 输出文件路径
            sheet_name: Sheet名称
            headers: 表头列表，None表示使用第一条数据的键

        Returns:
            输出文件路径
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            if not data:
                wb.save(output_path)
                return output_path

            # 确定表头
            if headers is None:
                headers = list(data[0].keys())

            # 写入表头
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(output_path)

            logger.info(
                "excel_write_success",
                output_path=output_path,
                sheet_name=sheet_name,
                rows_written=len(data),
                headers=headers
            )

            return output_path

        except Exception as e:
            logger.error(
                "excel_write_failed",
                output_path=output_path,
                error=str(e)
            )
            raise

    @staticmethod
    def check_file_exists(file_path: str) -> bool:
        """
        检查文件是否存在

        Args:
            file_path: 文件路径

        Returns:
            文件是否存在
        """
        return Path(file_path).exists() and Path(file_path).is_file()

    @staticmethod
    def validate_excel_file(file_path: str) -> tuple[bool, Optional[str]]:
        """
        验证Excel文件的有效性

        Args:
            file_path: 文件路径

        Returns:
            (是否有效, 错误信息)
        """
        # 检查文件是否存在
        if not ExcelService.check_file_exists(file_path):
            return False, f"文件不存在: {file_path}"

        # 检查文件扩展名
        if not file_path.endswith('.xlsx'):
            return False, f"文件格式错误，必须是.xlsx: {file_path}"

        # 尝试打开文件
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            wb.close()
            return True, None
        except Exception as e:
            return False, f"文件损坏或无法读取: {str(e)}"
