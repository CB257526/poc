"""Excel读写服务"""

from openpyxl import load_workbook, Workbook
from typing import List, Dict, Any, Optional
from pathlib import Path
import structlog

logger = structlog.get_logger()


class ExcelService:
    """Excel文件读写服务"""

    @staticmethod
    def read_sheet_as_dicts(
        file_path: str,
        sheet_name: Optional[str] = None,
        skip_rows: int = 0,
        max_rows: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """
        读取Excel sheet并转换为字典列表

        Args:
            file_path: Excel文件路径
            sheet_name: Sheet名称，None表示第一个sheet
            skip_rows: 跳过的行数（通常用于跳过标题前的说明）
            max_rows: 最多读取的行数，None表示全部读取

        Returns:
            字典列表，每个字典代表一行数据
        """
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            # 获取表头（第一行数据）
            headers = []
            header_row = skip_rows + 1
            for cell in ws[header_row]:
                headers.append(cell.value)

            # 读取数据
            rows = []
            start_row = skip_rows + 2
            end_row = ws.max_row + 1

            if max_rows:
                end_row = min(end_row, start_row + max_rows)

            for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
                # 跳过空行
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                row_dict = {}
                for header, value in zip(headers, row):
                    # 处理None值
                    if value is None:
                        row_dict[header] = None
                    else:
                        # 转换为字符串并去除首尾空格
                        row_dict[header] = str(value).strip() if value != "" else None

                rows.append(row_dict)

            wb.close()

            logger.info(
                "excel_read_success",
                file_path=file_path,
                sheet_name=sheet_name or ws.title,
                rows_read=len(rows),
                headers=headers
            )

            return rows

        except Exception as e:
            logger.error(
                "excel_read_failed",
                file_path=file_path,
                sheet_name=sheet_name,
                error=str(e)
            )
            raise

    @staticmethod
    def get_metadata(file_path: str) -> Dict[str, Any]:
        """
        获取Excel文件元信息

        Args:
            file_path: Excel文件路径

        Returns:
            包含文件大小、sheet名称、行列数等信息的字典
        """
        import os

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)

            metadata = {
                "file_path": file_path,
                "size_bytes": os.path.getsize(file_path),
                "sheet_names": wb.sheetnames,
                "sheets": {}
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                metadata["sheets"][sheet_name] = {
                    "max_row": ws.max_row,
                    "max_column": ws.max_column
                }

            wb.close()

            logger.info(
                "excel_metadata_extracted",
                file_path=file_path,
                sheets=len(wb.sheetnames)
            )

            return metadata

        except Exception as e:
            logger.error(
                "excel_metadata_failed",
                file_path=file_path,
                error=str(e)
            )
            raise

    @staticmethod
    def write_dicts_to_excel(
        data: List[Dict[str, Any]],
        output_path: str,
        sheet_name: str = "Sheet1",
        headers: Optional[List[str]] = None
    ) -> str:
        """
        将字典列表写入Excel文件

        Args:
            data: 数据列表
            output_path: 输出文件路径
            sheet_name: Sheet名称
            headers: 表头列表，None表示使用第一条数据的键

        Returns:
            输出文件路径
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            if not data:
                wb.save(output_path)
                return output_path

            # 确定表头
            if headers is None:
                headers = list(data[0].keys())

            # 写入表头
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(output_path)

            logger.info(
                "excel_write_success",
                output_path=output_path,
                sheet_name=sheet_name,
                rows_written=len(data),
                headers=headers
            )

            return output_path

        except Exception as e:
            logger.error(
                "excel_write_failed",
                output_path=output_path,
                error=str(e)
            )
            raise

    @staticmethod
    def check_file_exists(file_path: str) -> bool:
        """
        检查文件是否存在

        Args:
            file_path: 文件路径

        Returns:
            文件是否存在
        """
        return Path(file_path).exists() and Path(file_path).is_file()

    @staticmethod
    def validate_excel_file(file_path: str) -> tuple[bool, Optional[str]]:
        """
        验证Excel文件的有效性

        Args:
            file_path: 文件路径

        Returns:
            (是否有效, 错误信息)
        """
        # 检查文件是否存在
        if not ExcelService.check_file_exists(file_path):
            return False, f"文件不存在: {file_path}"

        # 检查文件扩展名
        if not file_path.endswith('.xlsx'):
            return False, f"文件格式错误，必须是.xlsx: {file_path}"

        # 尝试打开文件
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            wb.close()
            return True, None
        except Exception as e:
            return False, f"文件损坏或无法读取: {str(e)}"
