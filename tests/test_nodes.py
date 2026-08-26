"""测试新增业务节点"""

import pytest
from datetime import datetime
from workflows.models import WorkflowContext
from workflows.nodes.node_00_input import Node00Input
from workflows.nodes.node_02_fill_publication import Node02FillPublication
from workflows.nodes.node_03_match_media import Node03MatchMedia
from workflows.nodes.node_04_match_account import Node04MatchAccount
from workflows.nodes.node_05_calculate_fee import Node05CalculateFee
from workflows.nodes.node_06_generate_payment import Node06GeneratePayment


@pytest.fixture
def base_context():
    """创建基础上下文"""
    return WorkflowContext(
        run_id="test_run",
        run_started_at=datetime.now(),
        input_file="test.xlsx",
        table_dir="./table",
        records=[
            {
                "id": "1",
                "链接": "https://zhihu.com/article/123456",
                "媒体": "测试媒体",
                "平台": "知乎"
            }
        ]
    )


def test_node_03_normalize_name():
    """测试名称标准化"""
    node = Node03MatchMedia()

    # 测试移除空格
    assert node._normalize_name("测试 媒体") == "测试媒体"
    assert node._normalize_name("测试　媒体") == "测试媒体"

    # 测试转换小写
    assert node._normalize_name("ABC") == "abc"
    assert node._normalize_name("Test Media") == "testmedia"


def test_node_04_normalize_name():
    """测试账户名称标准化"""
    node = Node04MatchAccount()

    # 测试移除空格和转换小写
    assert node._normalize_name("测试 账户") == "测试账户"
    assert node._normalize_name("ABC DEF") == "abcdef"


def test_node_05_parse_fee():
    """测试费用解析"""
    node = Node05CalculateFee()

    # 测试数字
    assert node._parse_fee(100) == 100.0
    assert node._parse_fee(100.5) == 100.5

    # 测试字符串
    assert node._parse_fee("100") == 100.0
    assert node._parse_fee("100.5") == 100.5

    # 测试带货币符号
    assert node._parse_fee("¥100") == 100.0
    assert node._parse_fee("￥100.5") == 100.5

    # 测试带逗号
    assert node._parse_fee("1,000") == 1000.0
    assert node._parse_fee("1,000.50") == 1000.5

    # 测试None和无效值
    assert node._parse_fee(None) is None
    assert node._parse_fee("invalid") is None


def test_node_06_extract_month():
    """测试月份提取"""
    node = Node06GeneratePayment()

    # 测试标准日期格式
    assert node._extract_month("2024-08-15") == "2024-08"
    assert node._extract_month("2024/08/15") == "2024-08"
    assert node._extract_month("20240815") == "2024-08"

    # 测试只有年月
    assert node._extract_month("2024-08") == "2024-08"
    assert node._extract_month("2024/08") == "2024-08"

    # 测试datetime对象
    dt = datetime(2024, 8, 15)
    assert node._extract_month(dt) == "2024-08"

    # 测试无效输入
    assert node._extract_month(None) == ""
    assert node._extract_month("") == ""
    assert node._extract_month("invalid") == ""


def test_workflow_with_all_nodes(base_context):
    """测试完整工作流（集成测试，需要真实表格文件）"""
    # 这个测试需要真实的表格文件，标记为可能跳过
    import os

    # 检查表格文件是否存在
    table_files = [
        "./table/2-约稿资料.xlsx",
        "./table/3-媒体库.xlsx",
        "./table/4-账户信息.xlsx",
        "./table/5-费用.xlsx"
    ]

    if not all(os.path.exists(f) for f in table_files):
        pytest.skip("需要真实的表格文件来运行集成测试")

    # 如果文件存在，测试各节点能够正常初始化
    node2 = Node02FillPublication()
    assert node2.node_id == "node_02"
    assert node2.node_name == "完善发布信息"

    node3 = Node03MatchMedia()
    assert node3.node_id == "node_03"
    assert node3.node_name == "匹配媒体库"

    node4 = Node04MatchAccount()
    assert node4.node_id == "node_04"
    assert node4.node_name == "匹配账户信息"

    node5 = Node05CalculateFee()
    assert node5.node_id == "node_05"
    assert node5.node_name == "计算费用"

    node6 = Node06GeneratePayment()
    assert node6.node_id == "node_06"
    assert node6.node_name == "生成付款表"


def test_node_00_applies_frontend_media_name_correction(monkeypatch, tmp_path):
    """前端按原Excel行号提交修正后，节点0应使用新媒体名重新校验。"""
    input_file = tmp_path / "1-链接.xlsx"
    input_file.touch()
    for filename in ["3-媒体库.xlsx", "4-账户信息.xlsx", "5-费用.xlsx"]:
        (tmp_path / filename).touch()

    monkeypatch.setattr(
        "workflows.nodes.node_00_input.ExcelService.read_link_sheet",
        lambda *_args, **_kwargs: [{
            "主题": "主题1", "媒体": "Oxgen", "row_number": 2,
            "链接": ["https://example.com/1"],
        }],
    )
    monkeypatch.setattr(
        "workflows.nodes.node_00_input.ExcelService.read_sheet_as_dicts",
        lambda *_args, **_kwargs: [{"媒体名称": "Oxygen"}],
    )
    context = WorkflowContext(
        run_id="test-correction",
        input_file=str(input_file),
        table_dir=str(tmp_path),
        config={"media_name_corrections": {"2": "Oxygen"}},
    )

    output = Node00Input().process(context)

    assert output.data["records"][0]["媒体"] == "Oxygen"
    assert not any(issue.code == "MEDIA_NOT_IN_LIBRARY" for issue in output.issues)


def test_node_00_rejects_malformed_original_urls(monkeypatch, tmp_path):
    """表1 原始链接必须符合 URL 规范：ww、全角小数点等应报 INVALID_URL。"""
    input_file = tmp_path / "1-链接.xlsx"
    input_file.touch()
    for filename in ["3-媒体库.xlsx", "4-账户信息.xlsx", "5-费用.xlsx"]:
        (tmp_path / filename).touch()

    monkeypatch.setattr(
        "workflows.nodes.node_00_input.ExcelService.read_link_sheet",
        lambda *_args, **_kwargs: [{
            "主题": "主题1",
            "媒体": "Oxygen",
            "row_number": 2,
            "链接": [
                "知乎：https://ww.zhihu.com/zvideo/123",
                "微博：https://weibo。com/123",
                "B站：https://www.bilibili.com/video/BV1RgzpBHEQL/",
            ],
        }],
    )
    monkeypatch.setattr(
        "workflows.nodes.node_00_input.ExcelService.read_sheet_as_dicts",
        lambda *_args, **_kwargs: [{"媒体名称": "Oxygen"}],
    )
    context = WorkflowContext(
        run_id="test-invalid-url",
        input_file=str(input_file),
        table_dir=str(tmp_path),
    )

    output = Node00Input().process(context)

    invalid = [issue for issue in output.issues if issue.code == "INVALID_URL"]
    assert len(invalid) == 2
    assert any("ww.zhihu.com" in issue.message for issue in invalid)
    assert any("weibo" in issue.message for issue in invalid)
    assert not any("bilibili.com" in issue.message for issue in invalid)
    assert output.data["records"][0]["媒体"] == "Oxygen"


def test_node_00_accepts_valid_share_text(monkeypatch, tmp_path):
    """分享文案夹在 URL 后时，合法链接仍应通过。"""
    input_file = tmp_path / "1-链接.xlsx"
    input_file.touch()
    for filename in ["3-媒体库.xlsx", "4-账户信息.xlsx", "5-费用.xlsx"]:
        (tmp_path / filename).touch()

    monkeypatch.setattr(
        "workflows.nodes.node_00_input.ExcelService.read_link_sheet",
        lambda *_args, **_kwargs: [{
            "主题": "主题1",
            "媒体": "Oxygen",
            "row_number": 2,
            "链接": [
                'https://v.kuaishou.com/n8UrbDV9 少女专场 !"COS "鸣潮 "原神',
                "知乎：https://www.zhihu.com/zvideo/1997648866380632485",
            ],
        }],
    )
    monkeypatch.setattr(
        "workflows.nodes.node_00_input.ExcelService.read_sheet_as_dicts",
        lambda *_args, **_kwargs: [{"媒体名称": "Oxygen"}],
    )
    context = WorkflowContext(
        run_id="test-valid-url",
        input_file=str(input_file),
        table_dir=str(tmp_path),
    )

    output = Node00Input().process(context)

    assert not any(issue.code == "INVALID_URL" for issue in output.issues)


# ============================================================
# 回归测试：针对真实表格结构修复的字段映射与转换逻辑
# ============================================================

def test_node_03_media_field_mapping():
    """测试媒体库列名映射（真实表头：媒体级别/粉丝量，而非媒体等级/粉丝数）"""
    node = Node03MatchMedia()
    node._media_data = [
        {"序号": 1, "媒体名称": "Alex Cui", "媒体级别": "FA", "粉丝量": "20.4w（微博129.2w）"},
        {"序号": 2, "媒体名称": "Oxygen", "媒体级别": "FC", "粉丝量": "4.3w"},
    ]

    matched = node._match_media_info("Alex Cui")
    assert matched is not None
    # 关键：能取到媒体级别和粉丝量
    assert matched.get("媒体级别") == "FA"
    assert matched.get("粉丝量") == "20.4w（微博129.2w）"


def test_node_03_rejects_duplicate_media_name():
    """媒体库重名时不得静默使用第一条记录。"""
    node = Node03MatchMedia()
    node._media_data = [
        {"媒体名称": "Alex Cui", "媒体级别": "FA", "粉丝量": "20.4w"},
        {"媒体名称": "Alex Cui", "媒体级别": "FB", "粉丝量": "10w"},
    ]

    assert node._match_media_info("Alex Cui") is None
    assert len(node._build_media_index()["alexcui"]) == 2


def test_node_03_process_success(monkeypatch, base_context):
    """匹配成功时写入媒体等级、粉丝量及可处理状态。"""
    monkeypatch.setattr(
        "workflows.nodes.node_03_match_media.ExcelService.read_sheet_as_dicts",
        lambda _: [{"媒体名称": "测试媒体", "媒体级别": "FA", "粉丝量": "20w"}],
    )

    output = Node03MatchMedia().process(base_context)

    assert output.success is True
    assert output.metrics.success_count == 1
    assert output.metrics.error_count == 0
    assert base_context.records[0]["媒体等级"] == "FA"
    assert base_context.records[0]["粉丝量"] == "20w"
    assert base_context.records[0]["media_match_status"] == "matched"
    assert base_context.records[0]["processable"] is True


def test_node_03_process_media_not_found_is_error(monkeypatch, base_context):
    """表1媒体名无法匹配时生成error并阻止后续业务处理。"""
    monkeypatch.setattr(
        "workflows.nodes.node_03_match_media.ExcelService.read_sheet_as_dicts",
        lambda _: [{"媒体名称": "另一个媒体", "媒体级别": "FA", "粉丝量": "20w"}],
    )

    output = Node03MatchMedia().process(base_context)

    assert output.metrics.error_count == 1
    assert any(issue.level == "error" and issue.code == "MEDIA_NOT_FOUND" for issue in output.issues)
    assert base_context.records[0]["media_match_status"] == "pending_confirmation"
    assert base_context.records[0]["processable"] is False


def test_node_03_does_not_duplicate_node_00_media_error(monkeypatch, base_context):
    """节点0已报告媒体名称错误时，节点3只设置拦截状态，不重复报错。"""
    base_context.add_issue(
        level="error",
        code="MEDIA_NOT_IN_LIBRARY",
        message="媒体库中没有这个媒体: 测试媒体",
        node_id="node_00",
        record_id="1",
    )
    monkeypatch.setattr(
        "workflows.nodes.node_03_match_media.ExcelService.read_sheet_as_dicts",
        lambda _: [{"媒体名称": "另一个媒体", "媒体级别": "FA", "粉丝量": "20w"}],
    )

    output = Node03MatchMedia().process(base_context)

    assert not any(issue.code == "MEDIA_NOT_FOUND" for issue in output.issues)
    assert base_context.records[0]["media_match_status"] == "pending_confirmation"
    assert base_context.records[0]["processable"] is False


def test_node_03_process_missing_fan_count_is_error(monkeypatch, base_context):
    """媒体库缺少粉丝量时生成error并标记记录不完整。"""
    monkeypatch.setattr(
        "workflows.nodes.node_03_match_media.ExcelService.read_sheet_as_dicts",
        lambda _: [{"媒体名称": "测试媒体", "媒体级别": "FA", "粉丝量": None}],
    )

    output = Node03MatchMedia().process(base_context)

    assert output.metrics.error_count == 1
    assert any(issue.level == "error" and issue.code == "MISSING_FAN_COUNT" for issue in output.issues)
    assert base_context.records[0]["media_match_status"] == "incomplete"
    assert base_context.records[0]["processable"] is False


def test_node_03_process_missing_media_level_is_error(monkeypatch, base_context):
    """媒体库缺少媒体级别时生成error并阻止后续业务处理。"""
    monkeypatch.setattr(
        "workflows.nodes.node_03_match_media.ExcelService.read_sheet_as_dicts",
        lambda _: [{"媒体名称": "测试媒体", "媒体级别": None, "粉丝量": "20w"}],
    )

    output = Node03MatchMedia().process(base_context)

    assert output.metrics.error_count == 1
    assert any(issue.level == "error" and issue.code == "MISSING_MEDIA_LEVEL" for issue in output.issues)
    assert base_context.records[0]["media_match_status"] == "incomplete"
    assert base_context.records[0]["processable"] is False


def test_node_03_process_duplicate_media_name_is_error(monkeypatch, base_context):
    """媒体库出现同名记录时生成error，不自动选择其中一条。"""
    monkeypatch.setattr(
        "workflows.nodes.node_03_match_media.ExcelService.read_sheet_as_dicts",
        lambda _: [
            {"媒体名称": "测试媒体", "媒体级别": "FA", "粉丝量": "20w"},
            {"媒体名称": "测试媒体", "媒体级别": "FB", "粉丝量": "10w"},
        ],
    )

    output = Node03MatchMedia().process(base_context)

    assert output.metrics.error_count == 1
    assert any(issue.level == "error" and issue.code == "DUPLICATE_MEDIA_NAME" for issue in output.issues)
    assert "媒体等级" not in base_context.records[0]
    assert base_context.records[0]["media_match_status"] == "pending_confirmation"
    assert base_context.records[0]["processable"] is False


def test_node_04_account_field_mapping():
    """测试账户信息列名映射（真实表头：户名/银行卡账号/开户行信息）"""
    node = Node04MatchAccount()
    node._account_data = [
        {
            "媒体": "Oxygen",
            "户名": "徐浩轩",
            "银行卡账号": "6217921478762521",
            "电话": "18837128611",
            "开户行信息（具体到支行）": "浦发郑州经三路支行",
            "开户行所在城市": "郑州",
        },
    ]

    matched = node._match_account_info("Oxygen")
    assert matched is not None
    # 关键：能取到户名、账号、开户行
    assert matched.get("户名") == "徐浩轩"
    assert matched.get("银行卡账号") == "6217921478762521"
    assert matched.get("开户行信息（具体到支行）") == "浦发郑州经三路支行"


def test_node_04_process_success(monkeypatch, base_context):
    """账户匹配成功时补齐所有输出字段。"""
    monkeypatch.setattr(
        "workflows.nodes.node_04_match_account.ExcelService.read_sheet_as_dicts",
        lambda _: [{
            "媒体": "测试媒体",
            "户名": "测试用户",
            "身份证号": "110101199001011234",
            "银行卡账号": "6222000000000000",
            "电话": "13800000000",
            "开户行信息（具体到支行）": "测试银行北京支行",
            "开户行所在城市": "北京",
        }],
    )

    output = Node04MatchAccount().process(base_context)

    assert output.metrics.success_count == 1
    assert output.metrics.error_count == 0
    assert base_context.records[0]["收款方"] == "测试用户"
    assert base_context.records[0]["身份证"] == "110101199001011234"
    assert base_context.records[0]["开户行所在城市"] == "北京"
    assert base_context.records[0]["account_match_status"] == "matched"
    assert base_context.records[0]["processable"] is True


def test_node_04_skips_unprocessable_record(monkeypatch, base_context):
    """节点3未通过的记录不得进入账户匹配。"""
    base_context.records[0]["processable"] = False
    monkeypatch.setattr(
        "workflows.nodes.node_04_match_account.ExcelService.read_sheet_as_dicts",
        lambda _: [],
    )

    output = Node04MatchAccount().process(base_context)

    assert output.metrics.success_count == 0
    assert output.metrics.error_count == 0
    assert output.issues == []
    assert base_context.records[0]["account_match_status"] == "skipped"


def test_node_04_account_not_found_is_error(monkeypatch, base_context):
    """账户信息无法匹配时生成error并阻止后续处理。"""
    monkeypatch.setattr(
        "workflows.nodes.node_04_match_account.ExcelService.read_sheet_as_dicts",
        lambda _: [{"媒体": "另一个媒体"}],
    )

    output = Node04MatchAccount().process(base_context)

    assert output.metrics.error_count == 1
    assert any(issue.level == "error" and issue.code == "ACCOUNT_NOT_FOUND" for issue in output.issues)
    assert base_context.records[0]["account_match_status"] == "not_found"
    assert base_context.records[0]["processable"] is False


def test_node_04_missing_required_field_is_error(monkeypatch, base_context):
    """任一账户必填字段为空时生成error。"""
    monkeypatch.setattr(
        "workflows.nodes.node_04_match_account.ExcelService.read_sheet_as_dicts",
        lambda _: [{
            "媒体": "测试媒体",
            "户名": "测试用户",
            "身份证号": None,
            "银行卡账号": "6222000000000000",
            "电话": "13800000000",
            "开户行信息（具体到支行）": "测试银行北京支行",
            "开户行所在城市": "北京",
        }],
    )

    output = Node04MatchAccount().process(base_context)

    assert output.metrics.error_count == 1
    issue = next(issue for issue in output.issues if issue.code == "MISSING_ACCOUNT_FIELDS")
    assert issue.level == "error"
    assert issue.details["missing_fields"] == ["身份证"]
    assert base_context.records[0]["account_match_status"] == "incomplete"
    assert base_context.records[0]["processable"] is False


def test_node_04_duplicate_media_is_error(monkeypatch, base_context):
    """账户信息表中媒体重名时不得自动选择账户。"""
    monkeypatch.setattr(
        "workflows.nodes.node_04_match_account.ExcelService.read_sheet_as_dicts",
        lambda _: [
            {"媒体": "测试媒体", "户名": "用户A"},
            {"媒体": "测试媒体", "户名": "用户B"},
        ],
    )

    output = Node04MatchAccount().process(base_context)

    assert output.metrics.error_count == 1
    assert any(issue.level == "error" and issue.code == "DUPLICATE_ACCOUNT_MEDIA" for issue in output.issues)
    assert base_context.records[0]["account_match_status"] == "duplicate"
    assert base_context.records[0]["processable"] is False


def test_node_05_calculate_fee_by_type():
    """测试费用计算按文章类型分列（真实表头：等级/视频费用/图文费用）"""
    node = Node05CalculateFee()
    node._fee_rules = [
        {"等级": "FA", "视频费用": 2000, "图文费用": 1000},
        {"等级": "FB", "视频费用": 1800, "图文费用": 800},
        {"等级": "FC", "视频费用": 1500, "图文费用": 600},
    ]

    # 视频类型取视频费用列
    assert node._calculate_fee("FA", "视频") == 2000.0
    assert node._calculate_fee("FB", "视频") == 1800.0
    # 图文类型取图文费用列
    assert node._calculate_fee("FA", "图文") == 1000.0
    assert node._calculate_fee("FC", "图文") == 600.0
    # 未知等级
    assert node._calculate_fee("ZZ", "图文") is None


def test_node_05_display_platform_and_sync_text():
    """约稿表平台列始终用主链接平台；同步平台保留原始多行文本。"""
    node = Node05CalculateFee()

    few_sync = {
        "primary_platform": "知乎",
        "sync_links": [{"url": "https://weibo.com/1"}] * 6,
    }
    many_sync = {
        "primary_platform": "知乎",
        "sync_links": [{"url": f"https://weibo.com/{i}"} for i in range(7)],
    }
    kuaishou_primary = {
        "primary_platform": "快手",
        "sync_links": [{"url": f"https://weibo.com/{i}"} for i in range(10)],
    }
    assert node._display_platform(few_sync) == "知乎"
    assert node._display_platform(many_sync) == "知乎"
    assert node._display_platform(kuaishou_primary) == "快手"

    record = {
        "链接": ["知乎 https://zhihu.com/p/1", "微博 https://weibo.com/1"],
    }
    assert node._sync_platform_text(record) == "知乎 https://zhihu.com/p/1\n微博 https://weibo.com/1"


def test_node_05_quote_details_include_sample_fields(monkeypatch, base_context):
    """费用明细需带上表2所需的开户行、城市、基础金额、同步平台。"""
    base_context.records[0].update({
        "processable": True,
        "媒体等级": "FA",
        "文章类型": "视频",
        "发布形式": "原创",
        "primary_link": "https://zhihu.com/p/1",
        "primary_platform": "知乎",
        "发布日期": "2026-01-28",
        "标题": "测试标题",
        "粉丝量": "20w",
        "收款方": "测试用户",
        "开户行": "测试银行北京支行",
        "开户行所在城市": "北京",
        "账号": "6222000000000000",
        "联系方式": "13800000000",
        "身份证": "110101199001011234",
        "链接": ["知乎 https://zhihu.com/p/1", "微博 https://weibo.com/1"],
        "sync_links": [{"url": "https://weibo.com/1", "raw_text": "微博 https://weibo.com/1"}],
        "截图": "/tmp/shot.png",
    })
    monkeypatch.setattr(
        "workflows.nodes.node_05_calculate_fee.ExcelService.read_sheet_as_dicts",
        lambda _: [{"等级": "FA", "视频费用": 2000, "图文费用": 1000}],
    )

    output = Node05CalculateFee().process(base_context)

    assert output.success is True
    detail = base_context.quote_details["details"][0]
    assert detail["基础金额"] == 2000.0
    assert detail["开户行"] == "测试银行北京支行"
    assert detail["开户行所在城市"] == "北京"
    assert detail["平台"] == "知乎"
    assert "知乎 https://zhihu.com/p/1" in detail["同步平台"]
    assert detail["发布形式"] is None
    assert detail["截图"] == "/tmp/shot.png"
    assert detail["eligible_for_monthly_summary"] is True
    assert base_context.quote_details["excluded_count"] == 0


def test_node_05_excludes_unprocessable_records_from_quote_and_monthly_data(monkeypatch, base_context):
    """仍有错误或待确认的记录不得形成费用明细，也不得进入月度统计。"""
    base_context.records[0].update({
        "processable": False,
        "媒体等级": "FA",
        "文章类型": "视频",
        "发布日期": "2026-08-20",
        "media_match_status": "pending_confirmation",
    })
    monkeypatch.setattr(
        "workflows.nodes.node_05_calculate_fee.ExcelService.read_sheet_as_dicts",
        lambda _: [{"等级": "FA", "视频费用": 2000}],
    )

    output = Node05CalculateFee().process(base_context)

    assert output.success is True
    assert base_context.quote_details["details"] == []
    assert base_context.quote_details["total_count"] == 0
    assert base_context.quote_details["total_fee"] == 0
    assert base_context.quote_details["excluded_count"] == 1


def test_node_02_browser_failure_does_not_stop_financial_workflow(monkeypatch, base_context):
    """网页抓取工具不可用时只告警，不应终止后续费用流程。"""
    async def unavailable(_records):
        raise RuntimeError("browser unavailable")

    monkeypatch.setattr(
        "workflows.nodes.node_02_fill_publication.scrape_publications",
        unavailable,
    )
    base_context.records[0].update({
        "primary_platform": "微博",
        "primary_link": "https://weibo.com/example",
    })

    output = Node02FillPublication().process(base_context)

    assert output.success is True
    assert output.issues[0].level == "warning"
    assert output.issues[0].code == "SCRAPING_UNAVAILABLE"
    assert base_context.records[0]["文章类型"] is None
    assert base_context.records[0]["平台"] == "微博"


def test_node_06_monthly_summary_ignores_ineligible_details():
    node = Node06GeneratePayment()
    details = [
        {
            "媒体": "正常媒体", "发布日期": "2026-08-20", "费用": 2000,
            "eligible_for_monthly_summary": True,
        },
        {
            "媒体": "异常媒体", "发布日期": "2026-08-20", "费用": 99999,
            "eligible_for_monthly_summary": False,
        },
    ]

    summary = node._generate_monthly_summary(details)

    assert summary["2026-08"] == {"媒体数": 1, "文章数": 1, "总费用": 2000.0}


def test_node_06_payment_rows_preserve_first_seen_order():
    """付款行按约稿资料中户名首次出现顺序汇总。"""
    node = Node06GeneratePayment()
    details = [
        {"收款方": "崔诚靓", "费用": 2000, "账号": "1", "身份证": "a", "联系方式": "11", "开户行": "行1"},
        {"收款方": "徐浩轩", "费用": 600, "账号": "2", "身份证": "b", "联系方式": "22", "开户行": "行2"},
        {"收款方": "徐浩轩", "费用": 600, "账号": "2", "身份证": "b", "联系方式": "22", "开户行": "行2"},
        {"收款方": "任开瑞", "费用": 600, "账号": "3", "身份证": "c", "联系方式": "33", "开户行": "行3"},
        {"收款方": "任开瑞", "费用": 600, "账号": "3", "身份证": "c", "联系方式": "33", "开户行": "行3"},
    ]

    rows = node._generate_payment_rows(details)

    assert [row["户名"] for row in rows] == ["崔诚靓", "徐浩轩", "任开瑞"]
    assert rows[1]["基础服务费"] == 1200.0
    assert rows[2]["基础服务费"] == 1200.0
    assert rows[0]["备注"] == ""


def test_node_06_write_quote_excel_matches_sample_layout(tmp_path):
    """约稿资料写出约稿 + 约稿费用合计，字段与合计行对齐样表。"""
    from openpyxl import load_workbook
    from PIL import Image as PILImage

    shot = tmp_path / "shot.png"
    PILImage.new("RGB", (40, 30), color=(20, 80, 160)).save(shot)

    node = Node06GeneratePayment()
    details = [
        {
            "媒体": "Alex Cui", "媒体等级": "FA", "粉丝量": "20.4w",
            "发布形式": "原创", "文章类型": "视频", "平台": "知乎",
            "标题": "t1", "链接": "https://zhihu.com/1", "发布日期": "2026-01-28",
            "收款方": "崔诚靓", "身份证": "id1", "账号": "acc1", "联系方式": "tel1",
            "开户行": "行1", "开户行所在城市": "北京", "基础金额": 2000, "奖励金额": None,
            "同步平台": "知乎 https://zhihu.com/1\n微博 https://weibo.com/1",
            "截图": str(shot),
        },
        {
            "媒体": "Oxygen", "媒体等级": "FC", "粉丝量": "4.3w",
            "发布形式": "原创", "文章类型": "图文", "平台": "知乎",
            "标题": "t2", "链接": "https://zhihu.com/2", "发布日期": "2026-01-29",
            "收款方": "徐浩轩", "身份证": "id2", "账号": "acc2", "联系方式": "tel2",
            "开户行": "行2", "开户行所在城市": "郑州", "基础金额": 600, "奖励金额": None,
            "同步平台": "知乎 https://zhihu.com/2",
        },
        {
            "媒体": "Oxygen", "媒体等级": "FC", "粉丝量": "4.3w",
            "发布形式": "原创", "文章类型": "图文", "平台": "知乎",
            "标题": "t3", "链接": "https://zhihu.com/3", "发布日期": "2026-01-30",
            "收款方": "徐浩轩", "身份证": "id2", "账号": "acc2", "联系方式": "tel2",
            "开户行": "行2", "开户行所在城市": "郑州", "基础金额": 600, "奖励金额": None,
            "同步平台": "知乎 https://zhihu.com/3",
        },
    ]
    context = WorkflowContext(run_id="t", run_started_at=datetime.now(), input_file="t.xlsx")
    path = node._write_quote_detail_excel(context, details, output_dir=str(tmp_path))

    wb = load_workbook(path)
    assert wb.sheetnames == ["约稿", "约稿费用合计"]

    ws = wb["约稿"]
    assert [ws.cell(1, c).value for c in range(1, 21)] == [
        "媒体名称", "媒体级别", "粉丝量", "发布形式", "约稿类型",
        "平台", "标题", "发布链接", "作品截图", "发布日期",
        "约稿数量", "户名", "身份证", "账号", "电话",
        "开户行", "开户行所在城市", "基础金额", "奖励金额", "同步平台",
    ]
    assert ws.cell(2, 4).value is None
    assert ws.cell(2, 5).value == "视频"
    assert ws.cell(2, 6).value == "知乎"
    assert ws.cell(2, 16).value == "行1"
    assert ws.cell(2, 17).value == "北京"
    assert ws.cell(2, 18).value == 2000
    assert ws.cell(2, 20).value.startswith("知乎")
    assert ws.cell(2, 9).value is None
    assert len(ws._images) == 1
    assert ws.cell(5, 17).value == "合计"
    assert ws.cell(5, 18).value == "=SUM(R2:R4)"

    ws_sum = wb["约稿费用合计"]
    assert [ws_sum.cell(1, c).value for c in range(1, 15)] == [
        "媒体名称", "媒体级别", "发布形式", "约稿类型", "约稿数量",
        "户名", "身份证", "账号", "电话", "开户行",
        "开户行所在城市", "基础金额", "奖励金额", "合计费用",
    ]
    # 费用合计表中发布形式=视频/图文；约稿类型留空待业务人员填写
    assert ws_sum.cell(2, 3).value == "视频"
    assert ws_sum.cell(2, 4).value is None
    assert ws_sum.cell(3, 14).value == 1200.0
    assert ws_sum.cell(4, 14).value is None
    assert "C3:C4" not in {str(r) for r in ws_sum.merged_cells.ranges}
    assert any(str(r) == "N3:N4" for r in ws_sum.merged_cells.ranges)
    assert ws_sum.cell(5, 11).value == "合计"
    assert ws_sum.cell(5, 12).value == "=SUM(L2:L4)"
    assert ws_sum.cell(5, 14).value == "=SUM(N2:N4)"


def test_node_06_write_payment_excel_matches_yunzhanghu_template(tmp_path):
    """付款表写成云账户银行卡上传模板，而不是自定义三表。"""
    from openpyxl import load_workbook
    from workflows.nodes.node_06_generate_payment import PAYMENT_TEMPLATE_ID

    node = Node06GeneratePayment()
    context = WorkflowContext(run_id="t", run_started_at=datetime.now(), input_file="t.xlsx")
    rows = [
        {
            "账号": "6228480038110246271",
            "户名": "崔诚靓",
            "身份证": "210102199108076931",
            "电话": "15021791031",
            "基础服务费": 2000,
            "备注": "",
        },
        {
            "账号": "6217921478762521",
            "户名": "徐浩轩",
            "身份证": "410105200005130051",
            "电话": "18837128611",
            "基础服务费": 1200,
            "备注": "",
        },
    ]
    path = node._write_payment_excel(context, rows, output_dir=str(tmp_path))

    wb = load_workbook(path)
    assert wb.sheetnames == ["上传模板"]
    ws = wb["上传模板"]
    assert ws["A1"].value == PAYMENT_TEMPLATE_ID
    assert "B1:F1" in {str(r) for r in ws.merged_cells.ranges}
    assert ws["A3"].value.startswith("付款表_")
    assert ws["B3"].value == "=COUNTA(F5:F9420)"
    assert ws["C3"].value == "=SUM(F5:F9420)"
    assert ws.cell(4, 2).value == "收款账号(个人银行卡号,必填)"
    assert ws.cell(5, 2).value == "6228480038110246271"
    assert ws.cell(5, 3).value == "崔诚靓"
    assert ws.cell(5, 6).value == 2000
    assert ws.cell(6, 6).value == 1200
    assert "付款汇总" not in wb.sheetnames
    assert "约稿明细" not in wb.sheetnames
    assert "月度汇总" not in wb.sheetnames
