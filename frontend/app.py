from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from difflib import get_close_matches
import os
import time

import altair as alt
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


st.set_page_config(
    page_title="约稿平台",
    page_icon="✓",
    layout="wide",
    initial_sidebar_state="expanded",
)

API_BASE_URL = os.getenv("WORKFLOW_API_URL", "http://127.0.0.1:8000").rstrip("/")
API_TIMEOUT = int(os.getenv("WORKFLOW_API_TIMEOUT", "30"))


def api_request(method: str, path: str, **kwargs):
    try:
        response = requests.request(method, f"{API_BASE_URL}{path}", timeout=API_TIMEOUT, **kwargs)
        response.raise_for_status()
        return response
    except requests.RequestException as exc:
        detail = ""
        if getattr(exc, "response", None) is not None:
            try:
                detail = exc.response.json().get("detail", "")
            except Exception:
                detail = exc.response.text
        raise RuntimeError(detail or f"无法连接后端服务：{exc}") from exc


def backend_review_frame(payload: dict) -> pd.DataFrame:
    bad_ids = {
        issue.get("record_id")
        for issue in payload.get("issues", [])
        if issue.get("code") == "MEDIA_NOT_IN_LIBRARY"
    }
    rows = []
    for record in payload.get("records", []):
        status = "未匹配" if record.get("record_id") in bad_ids else "✓ 已匹配"
        rows.append({
            "原表行号": record.get("row_number"),
            "主题": record.get("topic") or "",
            "媒体名称": record.get("media_name") or "",
            "链接数量": record.get("link_count", 0),
            "链接预览": record.get("link_preview") or "",
            "校验状态": status,
            "建议修改为": "" if status == "✓ 已匹配" else suggest_media_name(record.get("media_name")),
        })
    return pd.DataFrame(rows)


def start_backend_run(task_id: str) -> None:
    api_request("POST", f"/api/v1/tasks/{task_id}/run")
    st.session_state.processing_stage = "running"
    st.session_state.backend_status = "running"
    st.session_state.processed = False


def completed_details() -> list[dict]:
    result = st.session_state.get("backend_result") or {}
    return (result.get("quote_summary") or {}).get("details") or []


def details_frame() -> pd.DataFrame:
    rows = []
    for detail in completed_details():
        rows.append({
            "媒体名称": detail.get("媒体"),
            "发布平台": detail.get("平台"),
            "类型": detail.get("文章类型"),
            "媒体等级": detail.get("媒体等级"),
            "粉丝量": detail.get("粉丝量"),
            "约稿数量": 1,
            "单价": float(detail.get("基础金额") or detail.get("费用") or 0),
            "金额": float(detail.get("费用") or 0),
            "状态": "完成",
            "标题": detail.get("标题"),
            "发布链接": detail.get("链接"),
            "发布日期": detail.get("发布日期"),
        })
    return pd.DataFrame(rows)


def monthly_backend_data() -> dict | None:
    try:
        return api_request("GET", "/api/v1/analytics/monthly").json()
    except Exception:
        return None

st.markdown(
    """
    <style>
    :root { --brand:#165DFF; --ink:#17233D; --muted:#667085; --line:#E8ECF3; }
    .stApp { background:#F5F7FA; color:var(--ink); }
    [data-testid="stSidebar"] { background:#102A56; }
    [data-testid="stSidebar"] * { color:#F5F8FF; }
    [data-testid="stSidebar"] .stRadio label { padding:7px 10px; border-radius:8px; }
    [data-testid="stMetric"] { background:white; border:1px solid var(--line); border-radius:12px; padding:18px; box-shadow:0 2px 8px #17233D0A; }
    .hero { background:linear-gradient(120deg,#123E82,#1677FF); padding:26px 30px; border-radius:16px; color:white; margin-bottom:20px; }
    .hero h1 { margin:0 0 7px; font-size:29px; }
    .hero p { margin:0; color:#DCEAFF; }
    .panel { background:white; border:1px solid var(--line); border-radius:12px; padding:20px; margin-bottom:14px; }
    .status { display:inline-block; padding:4px 10px; border-radius:20px; font-size:13px; font-weight:600; }
    .done { background:#E8F7EF; color:#12805C; }
    .warn { background:#FFF4E5; color:#B54708; }
    .flow { display:flex; gap:8px; align-items:stretch; flex-wrap:wrap; }
    .flow-step { flex:1; min-width:120px; background:#F7FAFF; border:1px solid #D7E5FF; border-radius:10px; padding:14px; }
    .flow-num { color:var(--brand); font-weight:700; }
    .file-row { display:flex; justify-content:space-between; padding:13px 3px; border-bottom:1px solid var(--line); }
    div.stButton > button, div.stDownloadButton > button { border-radius:8px; font-weight:600; }
    .st-key-start_auto_processing button {
        min-width:168px; height:44px; padding:0 22px;
        color:white !important; border:0 !important; border-radius:11px !important;
        background:linear-gradient(135deg,#246BFD 0%,#1554D1 100%) !important;
        box-shadow:0 6px 16px rgba(22,93,255,.22);
        transition:transform .16s ease, box-shadow .16s ease, filter .16s ease;
    }
    .st-key-start_auto_processing button:hover {
        transform:translateY(-1px); filter:brightness(1.04);
        box-shadow:0 9px 20px rgba(22,93,255,.28);
    }
    .st-key-start_auto_processing button:active {
        transform:translateY(0); box-shadow:0 4px 10px rgba(22,93,255,.2);
    }
    .st-key-start_auto_processing button p { color:white !important; font-weight:650; }
    </style>
    """,
    unsafe_allow_html=True,
)


DETAILS = pd.DataFrame(
    [
        ["36氪汽车", "知乎", "图文", "A", "500万", 12, 5000, 60000, "完成"],
        ["懂车帝", "小红书", "图文", "A", "800万", 10, 4800, 48000, "完成"],
        ["汽车之家", "微博", "视频", "B", "300万", 8, 5500, 44000, "完成"],
        ["车云网", "微信", "图文", "B", "180万", 9, 3200, 28800, "完成"],
        ["新能源观察", "抖音", "视频", "A", "260万", 5, 6800, 34000, "待确认"],
        ["智驾前沿", "知乎", "图文", "C", "90万", 7, 2400, 16800, "完成"],
    ],
    columns=["媒体名称", "发布平台", "类型", "媒体等级", "粉丝量", "约稿数量", "单价", "金额", "状态"],
)

MONTHLY_UPLOADS = pd.DataFrame(
    [
        ["2026-08-02", "批次 01", 88, 44000, 26400],
        ["2026-08-05", "批次 02", 102, 51000, 31600],
        ["2026-08-08", "批次 03", 86, 46000, 28800],
        ["2026-08-11", "批次 04", 112, 55000, 35200],
        ["2026-08-14", "批次 05", 96, 53000, 33400],
        ["2026-08-17", "批次 06", 108, 58000, 36000],
        ["2026-08-20", "本次上传", 256, 153600, 78000],
    ],
    columns=["上传日期", "处理批次", "约稿数量", "图文费用", "视频费用"],
)
MONTHLY_UPLOADS["费用总额"] = MONTHLY_UPLOADS["图文费用"] + MONTHLY_UPLOADS["视频费用"]
MONTHLY_UPLOADS["上传日期"] = pd.to_datetime(MONTHLY_UPLOADS["上传日期"])

EXCEPTIONS = pd.DataFrame(
    [
        ["费用核算结果", "两个子表费用不一致", "核对“约稿”与“约稿费用合计”的媒体费用及总费用", "待确认"],
    ],
    columns=["核验对象", "校验事项", "处理建议", "状态"],
)

PAYMENT_ROWS = [
    [None, "DEMO-BANK-0001", "演示收款人A", "DEMO-ID-0001", "DEMO-PHONE-0001", 60000, "POC演示数据"],
    [None, "DEMO-BANK-0002", "演示收款人B", "DEMO-ID-0002", "DEMO-PHONE-0002", 48000, "POC演示数据"],
    [None, "DEMO-BANK-0003", "演示收款人C", "DEMO-ID-0003", "DEMO-PHONE-0003", 44000, "POC演示数据"],
]

# Demo 中模拟后台 3-媒体库可选择的标准名称。正式接入时由
# “输入预检接口”的 allowed_media_names 字段返回，不在前端写死。
MEDIA_LIBRARY_NAMES = [
    "Alex Cui", "Johnny Durn", "Oxygen", "景行",
    "36氪汽车", "懂车帝", "汽车之家", "车云网", "新能源观察", "智驾前沿",
]


def normalize_media_name(value: object) -> str:
    return "".join(str(value or "").split()).lower()


def suggest_media_name(value: object) -> str:
    normalized_to_name = {normalize_media_name(name): name for name in MEDIA_LIBRARY_NAMES}
    matches = get_close_matches(normalize_media_name(value), list(normalized_to_name), n=1, cutoff=0.45)
    return normalized_to_name[matches[0]] if matches else "请从媒体库选择"


def parse_link_workbook(uploaded_file) -> pd.DataFrame:
    """按后端 ExcelService.read_link_sheet 的规则读取表1媒体块。"""
    workbook = load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True, read_only=True)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    records = []
    current_topic = None
    current_media = None
    media_row = None
    links = []

    def clean(value):
        text = str(value).strip() if value is not None else ""
        return text or None

    def flush_media():
        nonlocal current_media, media_row, links
        if current_media and links:
            records.append({
                "原表行号": media_row,
                "主题": current_topic or "",
                "媒体名称": current_media,
                "链接数量": len(links),
                "链接预览": "\n".join(links[:2]) + ("\n…" if len(links) > 2 else ""),
            })
        current_media, media_row, links = None, None, []

    try:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            first = clean(row[0] if len(row) >= 1 else None)
            second = clean(row[1] if len(row) >= 2 else None)
            if not second:
                flush_media()
                if first:
                    current_topic = first
                continue
            if first:
                flush_media()
                current_media, media_row = first, row_number
            if current_media:
                links.append(second)
        flush_media()
    finally:
        workbook.close()

    if not records:
        raise ValueError("没有读取到有效的媒体与链接，请检查表1是否符合两列分层模板。")
    return pd.DataFrame(records)


def demo_link_records() -> pd.DataFrame:
    """未上传文件时用于演示“媒体名修正后继续”的场景。"""
    return pd.DataFrame([
        [2, "主题1", "36氪汽车", 8, "https://example.com/36kr/1\n…"],
        [10, "主题1", "懂车帝", 6, "https://example.com/dcd/1\n…"],
        [16, "主题1", "汽车之加", 8, "https://example.com/autohome/1\n…"],
    ], columns=["原表行号", "主题", "媒体名称", "链接数量", "链接预览"])


def validate_media_records(frame: pd.DataFrame) -> pd.DataFrame:
    valid_names = {normalize_media_name(name) for name in MEDIA_LIBRARY_NAMES}
    checked = frame.copy()
    checked["校验状态"] = checked["媒体名称"].map(
        lambda name: "✓ 已匹配" if normalize_media_name(name) in valid_names else "未匹配"
    )
    checked["建议修改为"] = checked.apply(
        lambda row: "" if row["校验状态"] == "✓ 已匹配" else suggest_media_name(row["媒体名称"]),
        axis=1,
    )
    return checked


def workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def result_workbook_bytes() -> bytes:
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "约稿"
    detail_frame = DETAILS.rename(columns={"发布平台": "平台", "类型": "约稿类型"}).copy()
    detail_frame["发布形式"] = "原创"
    detail_frame["标题"] = "POC演示约稿内容"
    detail_frame["发布链接"] = "https://example.com/demo"
    detail_frame["作品截图"] = ""
    detail_frame["发布日期"] = "2026-08-18"
    detail_frame["基础金额"] = detail_frame["单价"]
    detail_frame["奖励金额"] = 0
    detail_frame["同步平台"] = "微博、小红书"
    detail_frame = detail_frame[[
        "媒体名称", "媒体等级", "粉丝量", "发布形式", "约稿类型", "平台", "标题", "发布链接",
        "作品截图", "发布日期", "约稿数量", "基础金额", "奖励金额", "金额", "同步平台", "状态"
    ]]
    for row in dataframe_to_rows(detail_frame, index=False, header=True):
        detail_sheet.append(row)

    summary_sheet = workbook.create_sheet("约稿费用合计")
    summary = DETAILS.groupby(["媒体名称", "媒体等级"], as_index=False).agg(
        约稿数量=("约稿数量", "sum"), 基础金额=("单价", "max"), 合计费用=("金额", "sum")
    )
    summary["发布形式"] = "原创"
    summary["约稿类型"] = "图文/视频"
    summary = summary[["媒体名称", "媒体等级", "发布形式", "约稿类型", "约稿数量", "基础金额", "合计费用"]]
    for row in dataframe_to_rows(summary, index=False, header=True):
        summary_sheet.append(row)

    header_fill = PatternFill("solid", fgColor="D9E8FF")
    thin = Side(style="thin", color="B8C4D6")
    for sheet in (detail_sheet, summary_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="17365D")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in column) + 2, 12), 28)
    return workbook_bytes(workbook)


def payment_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "上传模板"
    sheet.merge_cells("B1:F1")
    sheet["A1"] = "TEMPLATE-BANK-YZH006"
    sheet["B1"] = "自助签约发起付款表格模板-银行卡(单个批次文件最大支持10000条订单,单笔订单最大支持金额可在『业务中心-合作信息』中查看,请勿修改此条信息)"
    sheet.append(["批次号(与文件名称一致，必填)", "总笔数", "总金额/元(以显示金额汇总)"])
    sheet.append(["6-付款", "=COUNTA(F5:F9420)", "=SUM(F5:F9420)"])
    sheet.append([
        "平台企业订单号(非必填)", "收款账号(个人银行卡号,必填)", "收款户名(真实姓名,必填)",
        "身份证号(必填)", "联系电话(签约手机号,必填)", "基础服务费金额/元(四舍五入至分,必填)",
        "备注(36个字符以内,非必填)"
    ])
    for row in PAYMENT_ROWS:
        sheet.append(row)

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="8EA9C1")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number in (2, 4):
        for cell in sheet[row_number]:
            cell.font = Font(bold=True, color="17365D")
            cell.fill = light_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=1, max_row=4 + len(PAYMENT_ROWS), min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet["C3"].number_format = '¥#,##0.00'
    for row_number in range(5, 5 + len(PAYMENT_ROWS)):
        sheet[f"B{row_number}"].number_format = "@"
        sheet[f"D{row_number}"].number_format = "@"
        sheet[f"E{row_number}"].number_format = "@"
        sheet[f"F{row_number}"].number_format = "0.00"
    widths = {"A": 22.4, "B": 41.5, "C": 14.4, "D": 21.6, "E": 18.5, "F": 20, "G": 29.5}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A5"
    return workbook_bytes(workbook)


def output_files() -> dict[str, bytes]:
    return {
        "2-约稿资料_完成版.xlsx": result_workbook_bytes(),
        "6-付款.xlsx": payment_workbook_bytes(),
    }


def all_files_zip() -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for filename, content in output_files().items():
            archive.writestr(filename, content)
    return output.getvalue()


def hero(title: str, subtitle: str) -> None:
    st.markdown(f'<div class="hero"><h1>{title}</h1><p>{subtitle}</p></div>', unsafe_allow_html=True)


def cost_bar_chart(frame: pd.DataFrame, category: str, height: int = 330) -> None:
    chart = (
        alt.Chart(frame)
        .mark_bar(color="#165DFF", cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
        .encode(
            x=alt.X(f"{category}:N", title=None, axis=alt.Axis(labelAngle=0, labelLimit=120)),
            y=alt.Y("金额:Q", title="费用（元）", axis=alt.Axis(format=",.0f")),
            tooltip=[alt.Tooltip(f"{category}:N", title=category), alt.Tooltip("金额:Q", title="费用", format=",.0f")],
        )
        .properties(height=height)
    )
    st.altair_chart(chart, width="stretch")


if "processed" not in st.session_state:
    st.session_state.processed = False
if "processing_stage" not in st.session_state:
    st.session_state.processing_stage = "idle"
if "media_review_data" not in st.session_state:
    st.session_state.media_review_data = None
if "media_corrections" not in st.session_state:
    st.session_state.media_corrections = {}
if "uploaded_file_signature" not in st.session_state:
    st.session_state.uploaded_file_signature = None
if "backend_task_id" not in st.session_state:
    st.session_state.backend_task_id = None
if "backend_status" not in st.session_state:
    st.session_state.backend_status = None
if "backend_result" not in st.session_state:
    st.session_state.backend_result = None
if "backend_allowed_names" not in st.session_state:
    st.session_state.backend_allowed_names = []
if (
    "exception_data" not in st.session_state
    or list(st.session_state.exception_data.columns) != list(EXCEPTIONS.assign(修正内容="").columns)
    or len(st.session_state.exception_data) != len(EXCEPTIONS)
):
    st.session_state.exception_data = EXCEPTIONS.assign(修正内容="")
if "exception_editor_version" not in st.session_state:
    st.session_state.exception_editor_version = 0
if "exception_audit_message" not in st.session_state:
    st.session_state.exception_audit_message = ""
if "selected_exception" not in st.session_state:
    st.session_state.selected_exception = None

with st.sidebar:
    st.markdown("## 约稿平台")
    st.caption("POC · UI Demo v2")
    page = st.radio(
        "功能导航",
        ["首页概览", "数据处理", "约稿资料", "费用分析", "异常提醒", "文件输出"],
        label_visibility="collapsed",
    )
    st.divider()
    st.caption("演示数据状态")
    st.markdown('<span class="status done">● 数据正常</span>', unsafe_allow_html=True)
    st.caption("更新时间：2026-08-18 13:30")


if page == "首页概览":
    hero("约稿平台", "自动完成约稿资料整理、媒体信息匹配、费用计算及付款生成")
    live_details = details_frame()
    has_live_data = not live_details.empty
    live_total = float(live_details["金额"].sum()) if has_live_data else 231600
    live_media_count = int(live_details["媒体名称"].nunique()) if has_live_data else 128
    live_quote_count = len(live_details) if has_live_data else 256
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("任务状态", "已完成" if st.session_state.processed else "待处理", "本次文件 1 / 1" if has_live_data else "演示模式")
    c2.metric("媒体数量", f"{live_media_count}", "已完成匹配" if has_live_data else "演示数据")
    c3.metric("约稿数量", f"{live_quote_count}", "真实任务结果" if has_live_data else "图文 + 视频")
    c4.metric("费用总额", f"¥{live_total:,.0f}", "仅含校验通过数据")
    st.subheader("业务处理流程")
    steps = ["链接解析", "资料整理", "媒体匹配", "账户补全", "费用计算", "付款生成"]
    cards = "".join(f'<div class="flow-step"><span class="flow-num">✓ {i}</span><br>{step}</div>' for i, step in enumerate(steps, 1))
    st.markdown(f'<div class="flow">{cards}</div>', unsafe_allow_html=True)
    st.subheader("本次处理摘要")
    left, right = st.columns([1.5, 1])
    with left:
        st.markdown("**本次约稿类型分布**")
        if has_live_data:
            summary = live_details.groupby("类型", as_index=False).size().rename(columns={"类型": "约稿类型", "size": "约稿数量"})
        else:
            summary = pd.DataFrame([["图文", 168], ["视频", 88]], columns=["约稿类型", "约稿数量"])
        summary_chart = (
            alt.Chart(summary)
            .mark_bar(color="#165DFF", cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
            .encode(
                x=alt.X("约稿类型:N", title=None, axis=alt.Axis(labelAngle=0)),
                y=alt.Y("约稿数量:Q", title="约稿数量（条）"),
                tooltip=[
                    alt.Tooltip("约稿类型:N", title="约稿类型"),
                    alt.Tooltip("约稿数量:Q", title="约稿数量"),
                ],
            )
            .properties(height=230)
        )
        st.altair_chart(summary_chart, width="stretch")
    with right:
        if has_live_data:
            st.markdown('<div class="panel"><b>本次处理结果</b><br><br><span class="status done">全部通过</span><br><br>当前指标来自后端真实任务，异常记录未计入费用与月度统计。</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="panel"><b>费用一致性校验</b><br><br><span class="status warn">1 项待确认</span><br><br>“约稿”与“约稿费用合计”存在金额不一致，请核对后重新校验。</div>', unsafe_allow_html=True)

elif page == "数据处理":
    hero("数据处理", "上传本次链接表，系统自动生成约稿资料、费用汇总和付款文件")
    st.subheader("本次任务文件")
    st.caption("业务人员每次只需上传链接表")
    links_file = st.file_uploader("1-链接.xlsx", type=["xlsx"], key="task-links", help="本次约稿的全部发布及同步链接")
    current_signature = (links_file.name, links_file.size) if links_file else None
    if current_signature != st.session_state.uploaded_file_signature:
        st.session_state.uploaded_file_signature = current_signature
        st.session_state.processing_stage = "idle"
        st.session_state.media_review_data = None
        st.session_state.media_corrections = {}
        st.session_state.processed = False
        st.session_state.backend_task_id = None
        st.session_state.backend_status = None
        st.session_state.backend_result = None
        st.session_state.backend_allowed_names = []

    st.markdown("#### 基础配置状态")
    s1, s2, s3 = st.columns(3)
    s1.success("约稿资料模板 · 已配置")
    s2.success("媒体库 · 已配置")
    s3.success("账户信息 · 已配置")
    s4, s5 = st.columns(2)
    s4.success("费用规则 · 已配置")
    s5.success("付款模板 · 已配置")

    with st.expander("首次初始化或配置变更时更新", expanded=False):
        st.caption("以下文件由管理员在后台维护，业务人员日常不需要上传。2-约稿资料仅保留两个 Sheet 的空白表头。")
        config_cols = st.columns(2)
        config_names = ["2-约稿资料_空白模板.xlsx", "3-媒体库.xlsx", "4-账户信息.xlsx", "5-费用.xlsx", "6-付款模板.xlsx"]
        for i, name in enumerate(config_names):
            with config_cols[i % 2]:
                st.file_uploader(name, type=["xlsx"], key=f"config-{name}")
        st.button("保存基础配置", width="stretch")

    st.info("当前为 UI 演示版本：未上传文件时会使用一条媒体名有误的模拟数据，展示在线修正流程。")
    if st.button("开始自动处理  →", width="content", key="start_auto_processing"):
        try:
            if links_file:
                payload = api_request(
                    "POST",
                    "/api/v1/tasks/validate",
                    data=links_file.getvalue(),
                    headers={
                        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    },
                ).json()
                st.session_state.backend_task_id = payload["task_id"]
                st.session_state.backend_status = payload["status"]
                st.session_state.backend_allowed_names = payload.get("allowed_media_names", [])
                checked_rows = backend_review_frame(payload)
            else:
                checked_rows = validate_media_records(demo_link_records())
            st.session_state.media_review_data = checked_rows
            if (checked_rows["校验状态"] == "未匹配").any():
                st.session_state.processing_stage = "media_review"
                st.session_state.processed = False
            else:
                if st.session_state.backend_task_id:
                    start_backend_run(st.session_state.backend_task_id)
                else:
                    st.session_state.processing_stage = "completed"
                    st.session_state.processed = True
            st.rerun()
        except Exception as exc:
            st.error(f"处理失败：{exc}")

    if st.session_state.processing_stage == "media_review" and st.session_state.media_review_data is not None:
        review_data = st.session_state.media_review_data
        mismatch_count = int((review_data["校验状态"] == "未匹配").sum())
        st.error(f"输入预检暂停：发现 {mismatch_count} 个媒体名称无法匹配媒体库。请直接修改后再继续。")
        st.caption("系统尚未执行链接抓取和费用计算。主题、原表行号和链接信息仅用于定位，不能修改。")
        editor_options = st.session_state.backend_allowed_names or MEDIA_LIBRARY_NAMES
        edited_review = st.data_editor(
            review_data,
            width="stretch",
            hide_index=True,
            num_rows="fixed",
            key="media-name-review-editor",
            disabled=["原表行号", "主题", "链接数量", "链接预览", "校验状态", "建议修改为"],
            column_config={
                "媒体名称": st.column_config.SelectboxColumn(
                    "媒体名称（可修改）",
                    options=editor_options,
                    required=True,
                    help="请选择 3-媒体库中的标准媒体名称",
                ),
                "链接预览": st.column_config.TextColumn(width="large"),
                "校验状态": st.column_config.TextColumn(width="small"),
                "建议修改为": st.column_config.TextColumn(width="medium"),
            },
        )
        action_left, action_right = st.columns([2, 1])
        if action_left.button("重新校验并继续处理", type="primary", width="stretch"):
            try:
                corrections = {
                    str(row["原表行号"]): str(row["媒体名称"]).strip()
                    for _, row in edited_review.iterrows()
                }
                if st.session_state.backend_task_id:
                    payload = api_request(
                        "POST",
                        f"/api/v1/tasks/{st.session_state.backend_task_id}/corrections",
                        json={"media_name_corrections": corrections},
                    ).json()
                    rechecked = backend_review_frame(payload)
                else:
                    rechecked = validate_media_records(
                        edited_review.drop(columns=["校验状态", "建议修改为"], errors="ignore")
                    )
                st.session_state.media_review_data = rechecked
                remaining = rechecked[rechecked["校验状态"] == "未匹配"]
                if len(remaining):
                    st.warning(f"仍有 {len(remaining)} 个媒体名称未匹配，请继续修改。")
                else:
                    original_rows = review_data.set_index("原表行号")
                    corrected_rows = rechecked.set_index("原表行号")
                    st.session_state.media_corrections = {
                        str(row_number): corrected_rows.at[row_number, "媒体名称"]
                        for row_number in corrected_rows.index
                        if original_rows.at[row_number, "媒体名称"] != corrected_rows.at[row_number, "媒体名称"]
                    }
                    if st.session_state.backend_task_id:
                        start_backend_run(st.session_state.backend_task_id)
                    else:
                        st.session_state.processing_stage = "completed"
                        st.session_state.processed = True
                    st.rerun()
            except Exception as exc:
                st.error(f"重新校验失败：{exc}")
        if action_right.button("取消本次处理", width="stretch"):
            st.session_state.processing_stage = "idle"
            st.session_state.media_review_data = None
            st.session_state.media_corrections = {}
            st.session_state.backend_task_id = None
            st.rerun()

    if st.session_state.processing_stage == "running" and st.session_state.backend_task_id:
        st.info("后端正在执行链接抓取、媒体与账户匹配、费用计算和文件生成。")
        try:
            result = api_request("GET", f"/api/v1/tasks/{st.session_state.backend_task_id}").json()
            st.session_state.backend_result = result
            st.session_state.backend_status = result["status"]
            completed_count = len(result.get("completed_nodes", []))
            st.progress(min(completed_count / 7, 1.0), text=f"已完成 {completed_count}/7 个处理节点")
            if result["status"] == "completed":
                st.session_state.processing_stage = "completed"
                st.session_state.processed = True
                st.rerun()
            elif result["status"] == "failed":
                st.session_state.processing_stage = "failed"
                st.rerun()
            else:
                time.sleep(2)
                st.rerun()
        except Exception as exc:
            st.error(f"查询任务状态失败：{exc}")

    if st.session_state.processing_stage == "failed":
        result = st.session_state.backend_result or {}
        st.error(f"后端处理失败：{result.get('error') or '请查看任务问题详情'}")

    if st.session_state.processing_stage == "completed":
        if st.session_state.media_corrections:
            corrected_text = "、".join(
                f"第 {row} 行 → {name}" for row, name in st.session_state.media_corrections.items()
            )
            st.success(f"媒体名称重新校验通过（{corrected_text}），节点 1—6 已继续执行。")
        else:
            st.success("输入预检通过，节点 1—6 已继续执行。")
        if st.session_state.backend_task_id:
            st.caption(f"后端任务 ID：{st.session_state.backend_task_id}")
    if st.session_state.processed:
        st.success("最近一次任务已完成，可前往约稿资料、费用分析和文件输出查看结果。")

elif page == "约稿资料":
    hero("约稿资料", "查看、筛选本次处理后的约稿明细")
    source_details = details_frame()
    using_live_details = not source_details.empty
    if not using_live_details:
        source_details = DETAILS.copy()
        st.info("当前尚无已完成的后端任务，以下为演示数据。")
    else:
        st.success("当前展示后端返回的真实约稿明细。")
    f1, f2, f3 = st.columns(3)
    level = f1.selectbox("媒体等级", ["全部"] + sorted(source_details["媒体等级"].dropna().unique().tolist()))
    platform = f2.selectbox("发布平台", ["全部"] + sorted(source_details["发布平台"].dropna().unique().tolist()))
    state = f3.selectbox("状态", ["全部"] + sorted(source_details["状态"].dropna().unique().tolist()))
    filtered = source_details.copy()
    for column, value in [("媒体等级", level), ("发布平台", platform), ("状态", state)]:
        if value != "全部":
            filtered = filtered[filtered[column] == value]
    st.dataframe(
        filtered,
        width="stretch",
        hide_index=True,
        column_config={"单价": st.column_config.NumberColumn(format="¥%d"), "金额": st.column_config.NumberColumn(format="¥%d")},
    )
    st.caption(f"当前显示 {len(filtered)} 条{'真实' if using_live_details else '演示'}约稿记录。")

elif page == "费用分析":
    hero("费用分析", "查看本次任务费用，以及系统历史记录形成的当月累计汇总")
    st.success("统计口径：仅累计已通过媒体、账户及费用校验的约稿记录；待修改、待确认或处理失败的数据不会计入当月汇总。")
    live_cost_details = details_frame()
    has_live_cost = not live_cost_details.empty
    current_cost_data = live_cost_details if has_live_cost else DETAILS.copy()
    current_total = float(current_cost_data["金额"].sum())
    text_total = float(current_cost_data.loc[current_cost_data["类型"].isin(["图文", "文章", "图文类"]), "金额"].sum())
    video_total = float(current_cost_data.loc[current_cost_data["类型"].isin(["视频", "视频类"]), "金额"].sum())
    st.subheader("本次费用概览")
    c1, c2, c3 = st.columns(3)
    c1.metric("本次总费用", f"¥{current_total:,.0f}")
    c2.metric("本次图文费用", f"¥{text_total:,.0f}", f"{text_total/current_total:.1%}" if current_total else "0%")
    c3.metric("本次视频费用", f"¥{video_total:,.0f}", f"{video_total/current_total:.1%}" if current_total else "0%")
    st.subheader("当月 TOP 媒体费用")
    st.caption("汇总系统当月保存的约稿记录，并按媒体费用从高到低展示。")
    analytics = monthly_backend_data()
    if analytics and analytics.get("top_media"):
        top = pd.DataFrame(analytics["top_media"]).rename(columns={"media": "媒体名称", "total_fee": "金额"})[["媒体名称", "金额"]]
    else:
        top = current_cost_data.groupby("媒体名称", as_index=False)["金额"].sum().nlargest(5, "金额")
    cost_bar_chart(top, "媒体名称")

    st.divider()
    st.subheader("当月费用汇总")
    if analytics and analytics.get("batch_count"):
        st.caption(f"当前展示后端持久化的 {analytics['month']} 月真实处理记录。")
        monthly_total = float(analytics["total_fee"])
        monthly_count = int(analytics["quote_count"])
        monthly_batches = int(analytics["batch_count"])
        monthly_average = float(analytics["average_batch_fee"])
        monthly_table = pd.DataFrame(analytics["batches"])
        monthly_table["上传日期"] = pd.to_datetime(monthly_table["processed_at"]).dt.strftime("%Y-%m-%d %H:%M")
        monthly_table["处理批次"] = [f"批次 {i+1:02d}" for i in range(len(monthly_table))]
        monthly_table = monthly_table.rename(columns={"quote_count": "约稿数量", "total_fee": "费用总额"})
    else:
        st.caption("尚无后端持久化的当月记录，当前展示模拟历史批次数据。")
        monthly_total = int(MONTHLY_UPLOADS["费用总额"].sum())
        monthly_count = int(MONTHLY_UPLOADS["约稿数量"].sum())
        monthly_batches = len(MONTHLY_UPLOADS)
        monthly_average = int(monthly_total / monthly_batches)
        monthly_table = MONTHLY_UPLOADS.copy()
        monthly_table["上传日期"] = monthly_table["上传日期"].dt.strftime("%Y-%m-%d")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("当月累计费用", f"¥{monthly_total:,.0f}")
    m2.metric("当月处理批次", f"{monthly_batches} 次")
    m3.metric("当月约稿数量", f"{monthly_count} 条")
    m4.metric("平均每批费用", f"¥{monthly_average:,.0f}")

    chart_data = monthly_table.copy()
    chart_data["上传日期"] = pd.to_datetime(chart_data["上传日期"])
    monthly_chart = (
        alt.Chart(chart_data)
        .mark_line(point=True, color="#165DFF", strokeWidth=3)
        .encode(
            x=alt.X("上传日期:T", title=None, axis=alt.Axis(format="%m-%d", labelAngle=0)),
            y=alt.Y("费用总额:Q", title="费用（元）", axis=alt.Axis(format=",.0f")),
            tooltip=[
                alt.Tooltip("上传日期:T", title="上传日期", format="%Y-%m-%d"),
                alt.Tooltip("处理批次:N", title="批次"),
                alt.Tooltip("约稿数量:Q", title="约稿数量"),
                alt.Tooltip("费用总额:Q", title="费用总额", format=",.0f"),
            ],
        )
        .properties(height=300)
    )
    st.altair_chart(monthly_chart, width="stretch")
    display_columns = [column for column in ["上传日期", "处理批次", "约稿数量", "图文费用", "视频费用", "费用总额"] if column in monthly_table.columns]
    st.dataframe(
        monthly_table[display_columns],
        width="stretch",
        hide_index=True,
        column_config={
            "图文费用": st.column_config.NumberColumn(format="¥%d"),
            "视频费用": st.column_config.NumberColumn(format="¥%d"),
            "费用总额": st.column_config.NumberColumn(format="¥%d"),
        },
    )

elif page == "异常提醒":
    hero("费用一致性校验", "核对“约稿”与“约稿费用合计”两个子表的费用")
    exception_data = st.session_state.exception_data
    unresolved = exception_data[exception_data["状态"] != "已解决"]
    c1, c2, c3 = st.columns(3)
    c1.metric("待处理事项", str(len(unresolved)))
    c2.metric("待确认", str((exception_data["状态"] == "待确认").sum()))
    c3.metric("待校对", str((exception_data["状态"] == "待校对").sum()))

    st.subheader("待处理事项")
    st.caption("点击右侧状态按钮打开对应核验表；红色标记表示需要人工修改或确认的数据。")
    for row_index, row in exception_data.iterrows():
        media_col, issue_col, suggestion_col, status_col = st.columns([1.2, 1.7, 3, 1])
        media_col.markdown(f"**{row['核验对象']}**")
        issue_col.write(row["校验事项"])
        suggestion_col.caption(row["处理建议"])
        if row["状态"] == "已解决":
            status_col.success("已解决")
        elif status_col.button(row["状态"], key=f"open-exception-{row_index}", width="stretch"):
            st.session_state.selected_exception = row_index
            st.rerun()
        st.divider()

    selected_index = st.session_state.selected_exception
    if selected_index is not None and selected_index in exception_data.index:
        selected_row = exception_data.loc[selected_index]
        st.subheader(f"处理异常 · {selected_row['核验对象']}")
        st.error(f"需要处理：{selected_row['校验事项']}")
        correction = ""
        st.markdown("#### 2-约稿资料 · 两个子表费用核验")
        st.caption("系统已定位金额不一致的媒体。请先查看明细和计算依据，再确认汇总表中的红色金额。")

        st.markdown("##### ① 异常定位与计算依据")
        calculation_detail = pd.DataFrame(
            [
                ["汽车之家", "微博", "视频", 3, "B级视频核定价", 5500, 16500],
                ["汽车之家", "微博", "视频", 2, "B级视频核定价", 5500, 11000],
                ["汽车之家", "微博", "视频", 3, "B级视频核定价", 5500, 16500],
            ],
            columns=["媒体名称", "发布平台", "内容类型", "作品数量", "匹配费用规则", "核定单价", "应计费用"],
        )
        st.dataframe(
            calculation_detail,
            width="stretch",
            hide_index=True,
            column_config={
                "核定单价": st.column_config.NumberColumn(format="¥%d"),
                "应计费用": st.column_config.NumberColumn(format="¥%d"),
            },
        )
        rule_total = int(calculation_detail["应计费用"].sum())
        st.info(f"系统计算依据：汽车之家共 8 条视频 × 核定单价 ¥5,500 = 应计费用 ¥{rule_total:,.0f}。")

        st.markdown("##### ② 两个子表逐媒体对照")
        detail_cost = pd.DataFrame([
            ["36氪汽车", 60000], ["懂车帝", 48000], ["汽车之家", rule_total], ["车云网", 28800]
        ], columns=["媒体名称", "约稿子表费用"])
        summary_cost = pd.DataFrame([
            ["36氪汽车", 60000, "一致"], ["懂车帝", 48000, "一致"],
            ["汽车之家", 43000, "🔴 不一致"], ["车云网", 28800, "一致"]
        ], columns=["媒体名称", "🔴 约稿费用合计", "核验状态"])
        fee_c1, fee_c2 = st.columns(2)
        with fee_c1:
            st.markdown("**子表：约稿**")
            st.dataframe(detail_cost, width="stretch", hide_index=True)
            st.metric("约稿总费用", f"¥{detail_cost['约稿子表费用'].sum():,.0f}")
        with fee_c2:
            st.markdown("**子表：约稿费用合计**")
            summary_edited = st.data_editor(
                summary_cost, width="stretch", hide_index=True, num_rows="fixed",
                disabled=["媒体名称", "核验状态"], key=f"summary-check-{selected_index}",
                column_config={"🔴 约稿费用合计": st.column_config.NumberColumn(format="¥%d", min_value=0)},
            )
            st.metric("费用合计总额", f"¥{summary_edited['🔴 约稿费用合计'].sum():,.0f}")
        merged_cost = detail_cost.merge(summary_edited, on="媒体名称")
        merged_cost["差额"] = merged_cost["🔴 约稿费用合计"] - merged_cost["约稿子表费用"]
        amounts_match = (merged_cost["约稿子表费用"] == merged_cost["🔴 约稿费用合计"]).all()
        totals_match = detail_cost["约稿子表费用"].sum() == summary_edited["🔴 约稿费用合计"].sum()
        mismatch_rows = merged_cost[merged_cost["差额"] != 0]
        if len(mismatch_rows):
            mismatch = mismatch_rows.iloc[0]
            st.warning(
                f"当前差异：{mismatch['媒体名称']}在“约稿费用合计”中少计 "
                f"¥{abs(int(mismatch['差额'])):,.0f}；建议依据上方明细将金额确认至 ¥{int(mismatch['约稿子表费用']):,.0f}。"
            )
        else:
            st.success("两个子表的逐媒体金额和总费用现已一致，可以保存核验结果。")
        if amounts_match and totals_match:
            correction = f"两个子表费用已一致，总费用 ¥{detail_cost['约稿子表费用'].sum():,.0f}"

        save_c1, save_c2 = st.columns(2)
        saved = save_c1.button("保存核验结果", type="primary", width="stretch", key=f"save-check-{selected_index}")
        cancelled = save_c2.button("返回异常列表", width="stretch", key=f"cancel-check-{selected_index}")

        if saved:
            if not str(correction or "").strip():
                st.warning("请先修改红色金额，确保两个子表中各媒体费用及总费用全部一致。")
            else:
                st.session_state.exception_data.at[selected_index, "修正内容"] = correction
                st.session_state.exception_data.at[selected_index, "状态"] = "待校对"
                st.session_state.selected_exception = None
                st.session_state.exception_audit_message = "修改已保存，请点击“重新校对”执行规则验证。"
                st.rerun()
        if cancelled:
            st.session_state.selected_exception = None
            st.rerun()

    if st.button("重新校对", type="primary", width="stretch"):
        checked = st.session_state.exception_data.copy()
        for row_index, row in checked.iterrows():
            correction = str(row.get("修正内容", "") or "").strip()
            if row["状态"] == "待校对" and correction:
                checked.at[row_index, "状态"] = "已解决"
        resolved_count = int((checked["状态"] == "已解决").sum())
        st.session_state.exception_data = checked
        st.session_state.exception_audit_message = f"重新校对完成：{resolved_count} 项已通过，{len(checked) - resolved_count} 项仍需处理。"
        st.rerun()

    if st.session_state.exception_audit_message:
        if len(unresolved) == 0:
            st.success(st.session_state.exception_audit_message)
        else:
            st.info(st.session_state.exception_audit_message)
    st.warning("演示说明：异常中心仅展示两个子表的费用一致性校验，不对后台账户信息重复核验。")

else:
    hero("文件输出", "下载处理结果文件或一次性打包全部结果")
    st.subheader("处理结果文件")
    backend_files = (st.session_state.backend_result or {}).get("files", [])
    if st.session_state.backend_task_id and st.session_state.backend_status == "completed" and backend_files:
        names = {"quote_detail": "2-约稿资料_完成版.xlsx", "payment": "6-付款.xlsx"}
        for file_key in backend_files:
            filename = names.get(file_key, f"{file_key}.xlsx")
            try:
                content = api_request(
                    "GET", f"/api/v1/tasks/{st.session_state.backend_task_id}/files/{file_key}"
                ).content
                name_col, action_col = st.columns([5, 1])
                name_col.markdown(f'<div class="file-row"><span>✓ &nbsp; {filename}</span><span class="status done">后端已生成</span></div>', unsafe_allow_html=True)
                action_col.download_button("下载", content, filename, key=f"backend-download-{file_key}", width="stretch")
            except Exception as exc:
                st.error(f"读取 {filename} 失败：{exc}")
    else:
        st.caption("当前显示演示结果；上传真实文件并等待后端任务完成后，将自动切换为真实文件。")
        for filename, content in output_files().items():
            name_col, action_col = st.columns([5, 1])
            name_col.markdown(f'<div class="file-row"><span>✓ &nbsp; {filename}</span><span class="status done">演示文件</span></div>', unsafe_allow_html=True)
            action_col.download_button("下载", content, filename, key=f"download-{filename}", width="stretch")
        st.divider()
        st.download_button("下载全部演示文件（ZIP）", all_files_zip(), "约稿费用验收_处理结果.zip", type="primary", width="stretch")

st.caption("约稿平台 · 上传真实文件时调用后端工作流；未上传时使用模拟数据演示")
